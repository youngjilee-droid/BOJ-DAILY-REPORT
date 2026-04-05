import io
import json
import requests
import openpyxl
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta

try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

# ══════════════════════════════════════════════════════════════
# 페이지 설정 & 공통 스타일
# ══════════════════════════════════════════════════════════════
st.set_page_config(page_title="BOJ 광고 대시보드", layout="wide", page_icon="📊")
st.markdown("""
<style>
  .block-container { padding-top: 1.5rem; }
  .stTextArea textarea { font-family: 'Malgun Gothic', sans-serif; font-size: 14px; line-height: 1.7; }
  .section-title {
    font-size: 13px; font-weight: 600;
    background: #f0f2f6; border-left: 3px solid #4A90D9;
    padding: 6px 12px; border-radius: 0 6px 6px 0; margin: 20px 0 10px 0;
  }
  .landing-header {
    font-size: 12px; font-weight: 600; color: #fff;
    background: #4A90D9; padding: 4px 10px; border-radius: 4px;
    display: inline-block; margin: 12px 0 6px 0;
  }
  div[data-testid="stDataFrame"] { border: 1px solid #e8e8e8; border-radius: 8px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# 공통 상수
# ══════════════════════════════════════════════════════════════

LANDING_ALL = "전체"

COMMENT_MEDIA_MAP = {
    "Meta":          {"campaigns": ["Meta"],           "label": "Meta"},
    "TikTok":        {"campaigns": ["TikTok"],          "label": "TikTok"},
    "Kakao":         {"campaigns": ["Kakao", "카카오"], "label": "Kakao"},
    "Criteo":        {"campaigns": ["Criteo"],          "label": "Criteo"},
    "Buzzvil":       {"campaigns": ["Buzzvil"],         "label": "Buzzvil"},
    "Naver SSA":     {"campaigns": ["SSA"],             "label": "Naver SSA"},
    "Naver BSA":     {"campaigns": ["BSA"],             "label": "Naver BSA"},
    "Naver ADVoost": {"campaigns": ["ADVoost"],         "label": "Naver ADVoost"},
    "Push":          {"campaigns": ["Push"],            "label": "Push"},
    "전체":           {"campaigns": [],                 "label": "전체"},
}

# 리포트 표준 컬럼 순서 (Total_Raw 기준)
REPORT_COLS = ["날짜", "캠페인명", "광고그룹명", "광고명",
               "비용", "노출", "클릭", "구매", "매출액", "장바구니",
               "도달", "참여", "팔로우", "동영상 조회"]

# 대시보드용 표준 컬럼
STD_COLS = {
    "date":       "날짜",
    "campaign":   "캠페인명",
    "adgroup":    "광고그룹명",
    "ad":         "광고명/소재명",
    "spend":      "비용",
    "impression": "노출",
    "click":      "클릭",
    "purchase":   "구매",
    "revenue":    "매출액",
    "cart":       "장바구니",
    "landing":    "랜딩페이지",
}

# ══════════════════════════════════════════════════════════════
# RAW 변환 로직 — 매체별 transform 함수
# 각 함수는 RAW DataFrame → 표준 REPORT_COLS DataFrame 반환
# 비용은 모두 VAT 제외 처리 (총비용 / 1.1) 여부를 매체별로 명시
# ══════════════════════════════════════════════════════════════

def _to_int(v, default=0):
    try:
        return int(float(str(v).replace(",", "")))
    except:
        return default

def _to_float(v, default=0.0):
    try:
        return round(float(str(v).replace(",", "")), 2)
    except:
        return default

def _parse_advoost_date(s):
    """
    ADVoost '기간' 컬럼 → 날짜 파싱
    - '2026.03.28.'  → 2026-03-28
    - '2026-03-28'   → 2026-03-28 (이미 표준 형식)
    - '2026/03/28'   → 2026-03-28
    - '2026.3.28'    → 2026-03-28
    - None / NaN     → NaT
    """
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return pd.NaT
    s = str(s).strip().rstrip(".")
    if len(s) == 10 and s[4] == "-":
        return pd.to_datetime(s, errors="coerce")
    parts = s.replace("/", ".").split(".")
    if len(parts) == 3:
        return pd.to_datetime(f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}", errors="coerce")
    return pd.to_datetime(s, errors="coerce")

def transform_advoost(df_raw):
    """
    Naver ADVoost RAW → 표준 리포트 컬럼
    - 날짜   : 기간 → 날짜 (다양한 형식 자동 파싱)
    - 캠페인명: 캠페인 이름
    - 광고그룹명: 애셋 그룹 이름
    - 광고명  : 'Asset_all_items' 고정
    - 비용   : 총 비용 / 1.1 (VAT 제외)
    - 노출   : 노출
    - 클릭   : 클릭
    - 구매   : 구매완료수
    - 매출액  : 구매완료 전환 매출액
    - 장바구니: 장바구니 담기수
    - 도달~동영상 조회: None (ADVoost 미제공)
    """
    # 컬럼명 공백·특수문자 정규화 후 딕셔너리로 접근 (r.get 오작동 방지)
    df_raw = df_raw.copy()
    df_raw.columns = [str(c).strip() for c in df_raw.columns]

    # 컬럼 존재 여부 확인용 헬퍼
    def _col(row, name, default=None):
        return row[name] if name in row.index else default

    rows = []
    for _, r in df_raw.iterrows():
        rows.append({
            "날짜":       _parse_advoost_date(_col(r, "기간")),        # 기간 = 날짜
            "캠페인명":   str(_col(r, "캠페인 이름", "")).strip(),
            "광고그룹명": str(_col(r, "애셋 그룹 이름", "")).strip(),
            "광고명":     "Asset_all_items",
            "비용":       _to_float(_col(r, "총 비용", 0)) / 1.1,
            "노출":       _to_int(_col(r, "노출", 0)),
            "클릭":       _to_int(_col(r, "클릭", 0)),
            "구매":       _to_int(_col(r, "구매완료수", 0)),
            "매출액":     _to_int(_col(r, "구매완료 전환 매출액", 0)),
            "장바구니":   _to_int(_col(r, "장바구니 담기수", 0)),
            "도달":       None,
            "참여":       None,
            "팔로우":     None,
            "동영상 조회": None,
        })
    df = pd.DataFrame(rows, columns=REPORT_COLS)
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
    return df.sort_values("날짜", ascending=False).reset_index(drop=True)


def transform_meta(df_raw):
    """
    Meta RAW → 표준 리포트 컬럼
    전환 캠페인 / 참여 캠페인 두 유형 모두 처리 (컬럼 존재 여부로 자동 판별)

    [전환 캠페인 컬럼]
    - 구매      : 공유 항목이 포함된 구매              (없으면 0)
    - 매출액    : 공유 항목의 구매 전환값               (없으면 0)
    - 장바구니  : 공유 항목이 포함된 장바구니에 담기     (없으면 0)
    - 도달      : 도달                                (없으면 0)
    - 동영상 조회: 동영상 3초 이상 재생                 (없으면 0)

    [참여 캠페인 추가 컬럼]
    - 참여      : 게시물 참여          (컬럼 없으면 None)
    - 팔로우    : Instagram 팔로우    (컬럼 없으면 None)

    [공통]
    - 날짜      : 일 (이미 YYYY-MM-DD 형식)
    - 비용      : 지출 금액 (KRW) — VAT 별도 없음
    - 클릭      : 링크 클릭
    """
    # 컬럼명 공백·특수문자 정규화
    df_raw = df_raw.copy()
    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    cols = df_raw.columns.tolist()

    def _safe(r, key, as_int=True):
        """컬럼이 존재하면 값 변환, 없으면 None 반환"""
        if key not in r.index:
            return None
        val = r[key]
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return 0 if as_int else None
        return _to_int(val) if as_int else val

    def _col(r, key, default=""):
        """컬럼 값 안전하게 읽기"""
        return r[key] if key in r.index else default

    rows = []
    for _, r in df_raw.iterrows():
        rows.append({
            "날짜":       str(_col(r, "일")).strip(),
            "캠페인명":   str(_col(r, "캠페인 이름")).strip(),
            "광고그룹명": str(_col(r, "광고 세트 이름")).strip(),
            "광고명":     str(_col(r, "광고 이름")).strip(),
            "비용":       _to_float(_col(r, "지출 금액 (KRW)", 0)),
            "노출":       _to_int(_col(r, "노출", 0)),
            "클릭":       _to_int(_col(r, "링크 클릭", 0)),
            "구매":       _safe(r, "공유 항목이 포함된 구매")          or 0,
            "매출액":     _safe(r, "공유 항목의 구매 전환값")           or 0,
            "장바구니":   _safe(r, "공유 항목이 포함된 장바구니에 담기") or 0,
            "도달":       _safe(r, "도달")                            or 0,
            "참여":       _safe(r, "게시물 참여"),        # 컬럼 없으면 None
            "팔로우":     _safe(r, "Instagram 팔로우"),   # 컬럼 없으면 None
            "동영상 조회": _safe(r, "동영상 3초 이상 재생") or 0,
        })
    df = pd.DataFrame(rows, columns=REPORT_COLS)
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
    return df.sort_values(["날짜", "캠페인명", "광고그룹명", "광고명"], ascending=True).reset_index(drop=True)


# 매체명 → transform 함수 매핑
# 다른 매체 추가 시 transform 함수 + MEDIA_SIGNATURES 항목만 추가하면 됨
TRANSFORM_MAP = {
    "Naver ADVoost": transform_advoost,
    "Meta":          transform_meta,
    # "TikTok": transform_tiktok,
    # "Kakao":  transform_kakao,
}

# ── 매체 자동 감지용 고유 컬럼 시그니처 ──────────────────────
# 각 매체 RAW 파일에만 존재하는 컬럼명 집합 (1개 이상 일치 시 해당 매체로 판별)
# 우선순위: 리스트 앞쪽이 높음 (더 구체적인 매체를 먼저 배치)
MEDIA_SIGNATURES = [
    ("Naver ADVoost", ["애셋 그룹 이름", "총 비용", "구매완료수", "구매완료 전환 매출액"]),
    ("Meta",          ["지출 금액 (KRW)", "광고 세트 이름", "캠페인 이름"]),
    # 추후 추가 예시:
    # ("TikTok",   ["캠페인", "광고그룹", "완료된 결제"]),
    # ("Kakao",    ["청구금액", "광고그룹명", "구매금액"]),
    # ("Naver SSA",["캠페인명", "전환수", "전환매출"]),
]

def detect_media(df_raw):
    """
    RAW DataFrame의 컬럼명을 보고 매체를 자동 감지
    MEDIA_SIGNATURES의 고유 컬럼 중 가장 많이 일치하는 매체 반환
    일치 없으면 None 반환
    """
    cols = set(str(c).strip() for c in df_raw.columns)
    best_media, best_score = None, 0
    for media, sig_cols in MEDIA_SIGNATURES:
        score = sum(1 for c in sig_cols if c in cols)
        if score > best_score:
            best_media, best_score = media, score
    # 최소 1개 이상 일치해야 감지로 인정
    return best_media if best_score >= 1 else None

# ══════════════════════════════════════════════════════════════
# 대시보드용 매체별 RAW 컬럼 매핑 (기존 대시보드 뷰용)
# ══════════════════════════════════════════════════════════════
MEDIA_COL_MAP = {
    "Meta": {
        "date":       ["날짜", "Date", "일자", "기간"],
        "campaign":   ["캠페인 이름", "캠페인명", "Campaign name"],
        "adgroup":    ["광고 세트 이름", "광고세트명", "Ad set name"],
        "ad":         ["광고 이름", "광고명", "Ad name"],
        "spend":      ["금액 (KRW)", "비용", "Amount spent (KRW)", "Spend"],
        "impression": ["노출", "Impressions"],
        "click":      ["링크 클릭수", "클릭수", "Link clicks"],
        "purchase":   ["구매", "웹사이트 구매", "Purchases"],
        "revenue":    ["구매 전환값", "매출액", "Purchase conversion value"],
        "cart":       ["장바구니에 담기", "Add to cart"],
        "landing":    ["랜딩페이지", "Landing page"],
    },
    "TikTok": {
        "date":       ["날짜", "Date", "일자"],
        "campaign":   ["캠페인", "Campaign Name"],
        "adgroup":    ["광고그룹", "Ad Group Name"],
        "ad":         ["광고", "Ad Name"],
        "spend":      ["비용", "Cost", "Spend"],
        "impression": ["노출수", "Impressions"],
        "click":      ["클릭수", "Clicks"],
        "purchase":   ["완료된 결제", "구매", "Complete Payment"],
        "revenue":    ["완료된 결제 가치", "매출", "Complete Payment Value"],
        "cart":       ["장바구니 담기", "Add to Cart"],
        "landing":    ["랜딩페이지", "Landing page"],
    },
    "Kakao": {
        "date":       ["날짜", "일자", "기간"],
        "campaign":   ["캠페인", "캠페인명"],
        "adgroup":    ["광고그룹", "광고그룹명"],
        "ad":         ["소재", "소재명"],
        "spend":      ["비용", "청구금액"],
        "impression": ["노출수", "노출"],
        "click":      ["클릭수", "클릭"],
        "purchase":   ["구매수", "구매"],
        "revenue":    ["구매금액", "매출액"],
        "cart":       ["장바구니", "장바구니수"],
        "landing":    ["랜딩페이지", "Landing page"],
    },
    "Criteo": {
        "date":       ["날짜", "Date"],
        "campaign":   ["캠페인", "Campaign"],
        "adgroup":    ["광고세트", "Ad Set"],
        "ad":         ["광고", "Ad"],
        "spend":      ["비용", "Cost", "Spend"],
        "impression": ["노출", "Displays"],
        "click":      ["클릭", "Clicks"],
        "purchase":   ["구매", "Sales"],
        "revenue":    ["매출", "Revenue"],
        "cart":       ["장바구니", None],
        "landing":    ["랜딩페이지", None],
    },
    "Buzzvil": {
        "date":       ["날짜", "일자"],
        "campaign":   ["캠페인명"],
        "adgroup":    ["광고그룹명"],
        "ad":         ["소재명"],
        "spend":      ["비용"],
        "impression": ["노출"],
        "click":      ["클릭"],
        "purchase":   ["구매"],
        "revenue":    ["매출"],
        "cart":       [None],
        "landing":    ["랜딩페이지", None],
    },
    "Naver SSA": {
        "date":       ["날짜", "일자"],
        "campaign":   ["캠페인명"],
        "adgroup":    ["광고그룹명"],
        "ad":         ["광고명"],
        "spend":      ["비용", "광고비"],
        "impression": ["노출수", "노출"],
        "click":      ["클릭수", "클릭"],
        "purchase":   ["구매수", "전환수", "구매"],
        "revenue":    ["구매금액", "전환매출", "매출액"],
        "cart":       ["장바구니", None],
        "landing":    ["랜딩페이지", None],
    },
    "Naver BSA": {
        "date":       ["날짜", "일자"],
        "campaign":   ["캠페인명"],
        "adgroup":    ["광고그룹명"],
        "ad":         ["광고명"],
        "spend":      ["비용", "광고비"],
        "impression": ["노출수", "노출"],
        "click":      ["클릭수", "클릭"],
        "purchase":   ["구매수", "전환수"],
        "revenue":    ["구매금액", "전환매출"],
        "cart":       ["장바구니", None],
        "landing":    ["랜딩페이지", None],
    },
    "Naver ADVoost": {
        # 변환 함수(transform_advoost)가 있어서 컬럼 매핑은 변환 후 표준 컬럼 기준으로 동작
        "date":       ["날짜"],
        "campaign":   ["캠페인명"],
        "adgroup":    ["광고그룹명"],
        "ad":         ["광고명"],
        "spend":      ["비용"],
        "impression": ["노출"],
        "click":      ["클릭"],
        "purchase":   ["구매"],
        "revenue":    ["매출액"],
        "cart":       ["장바구니"],
        "landing":    ["랜딩페이지", None],
    },
    "Push": {
        "date":       ["날짜", "일자"],
        "campaign":   ["캠페인명"],
        "adgroup":    ["광고그룹명"],
        "ad":         ["광고명"],
        "spend":      ["비용"],
        "impression": ["노출"],
        "click":      ["클릭"],
        "purchase":   ["구매"],
        "revenue":    ["매출"],
        "cart":       [None],
        "landing":    ["랜딩페이지", None],
    },
}

DASHBOARD_MEDIA_LIST = list(MEDIA_COL_MAP.keys())

# ══════════════════════════════════════════════════════════════
# 세션 상태 초기화
# ══════════════════════════════════════════════════════════════
if "media_data"        not in st.session_state: st.session_state.media_data        = {}
if "media_warnings"    not in st.session_state: st.session_state.media_warnings    = {}
if "raw_reports"       not in st.session_state: st.session_state.raw_reports       = {}
if "converted_reports" not in st.session_state: st.session_state.converted_reports = {}
if "comment_history"   not in st.session_state: st.session_state.comment_history   = []
if "meta_api_df"       not in st.session_state: st.session_state.meta_api_df       = pd.DataFrame()
# raw_reports      : {매체명: DataFrame(표준 REPORT_COLS)}  ← RAW 변환 결과 저장
# converted_reports: {매체명: DataFrame(표준 REPORT_COLS)}  ← 다중 파일 취합 결과 저장
# meta_api_df      : Meta API로 수집한 데이터 (RAW 변환 페이지에서 활용)

# ══════════════════════════════════════════════════════════════
# Meta API — facebook Graph API v25.0 연동
# Streamlit Secrets: META_ACCESS_TOKEN, META_AD_ACCOUNT_ID
# ══════════════════════════════════════════════════════════════

_META_API_VERSION = "v25.0"
_META_BASE_URL    = f"https://graph.facebook.com/{_META_API_VERSION}"

def _meta_validate_date(date_text):
    try:
        datetime.strptime(date_text, "%Y-%m-%d")
        return date_text
    except Exception:
        raise ValueError(f"날짜 형식이 잘못되었습니다: {date_text}")

def _meta_normalize_account(ad_account_id):
    s = str(ad_account_id).strip()
    return s if s.startswith("act_") else f"act_{s}"

def _safe_int(v, default=0):
    try: return int(float(v))
    except: return default

def _safe_float(v, default=0.0):
    try: return float(v)
    except: return default

def _extract_action_total(action_list, target_types):
    if not isinstance(action_list, list): return 0.0
    total = 0.0
    for item in action_list:
        if str(item.get("action_type","")).strip() in target_types:
            total += _safe_float(item.get("value", 0))
    return total

def _extract_action_first(action_list, priority_types):
    """우선순위 순서대로 첫 번째 일치 값만 반환 (중복 합산 방지)"""
    if not isinstance(action_list, list): return 0.0
    action_map = {str(i.get("action_type","")).strip(): _safe_float(i.get("value",0))
                  for i in action_list}
    for t in priority_types:
        if t in action_map: return action_map[t]
    return 0.0

def _extract_action_first_fuzzy(action_list, priority_keywords):
    """키워드 포함 여부로 우선순위 매칭 (정확한 이름이 다를 때)"""
    if not isinstance(action_list, list): return 0.0
    for keywords in priority_keywords:
        for item in action_list:
            at = str(item.get("action_type","")).strip().lower()
            if all(kw.lower() in at for kw in keywords):
                return _safe_float(item.get("value", 0))
    return 0.0

def _meta_request(url, params):
    resp = requests.get(url, params=params, timeout=30)
    if resp.status_code != 200:
        raise RuntimeError(resp.text)
    result = resp.json()
    if "error" in result:
        raise RuntimeError(str(result["error"]))
    return result

def fetch_meta_data(start_date: str, end_date: str) -> pd.DataFrame:
    """
    Meta Graph API로 광고 인사이트 수집.
    Streamlit Secrets에 META_ACCESS_TOKEN, META_AD_ACCOUNT_ID 필요.
    반환: REPORT_COLS 기반 표준 DataFrame (매체='Meta')
    """
    try:
        start_date = _meta_validate_date(start_date)
        end_date   = _meta_validate_date(end_date)
    except Exception as e:
        st.error(str(e)); return pd.DataFrame()

    try:
        access_token   = str(st.secrets["META_ACCESS_TOKEN"]).strip()
        ad_account_id  = _meta_normalize_account(st.secrets["META_AD_ACCOUNT_ID"])
    except Exception:
        st.error("Streamlit Secrets에 META_ACCESS_TOKEN / META_AD_ACCOUNT_ID를 설정하세요.")
        return pd.DataFrame()

    url = f"{_META_BASE_URL}/{ad_account_id}/insights"

    collab_fields = [
        "date_start","date_stop","campaign_name","adset_name","ad_name",
        "impressions","clicks","spend","reach",
        "actions","catalog_segment_actions","catalog_segment_value",
    ]
    safe_fields = [
        "date_start","date_stop","campaign_name","adset_name","ad_name",
        "impressions","clicks","spend","reach","actions",
    ]

    base_params = {
        "access_token": access_token,
        "level": "ad",
        "limit": 500,
        "time_increment": 1,
        "action_attribution_windows": json.dumps(["7d_click","1d_view"]),
        "action_report_time": "conversion",
        "time_range": json.dumps({"since": start_date, "until": end_date}, ensure_ascii=False),
    }

    # 액션 타입 집합
    _link_click       = {"link_click"}
    _add_to_cart      = {"add_to_cart","omni_add_to_cart",
                         "offsite_conversion.fb_pixel_add_to_cart",
                         "onsite_web_add_to_cart","offsite_conversion.add_to_cart",
                         "onsite_conversion.add_to_cart"}
    _follow           = {"follow","follows","instagram_profile_follows","page_like","like"}
    _engagement       = {"page_engagement","post_engagement","post_interaction_gross","post"}
    _video_view       = {"video_view"}
    _initiate_checkout= {"initiate_checkout","omni_initiated_checkout",
                         "onsite_conversion.initiate_checkout","onsite_web_initiate_checkout"}
    _purchase_priority= ["shared_items_purchase","purchase_with_shared_items",
                         "purchases_with_shared_items","website_purchase_with_shared_items",
                         "catalog_segment_purchase","catalog_segment_omni_purchase",
                         "purchase","omni_purchase"]

    rows = []
    next_url = url
    next_params = dict(base_params)
    use_collab = True

    try:
        while next_url:
            params_copy = dict(next_params)
            if use_collab:
                try:
                    params_copy["fields"] = ",".join(collab_fields)
                    result = _meta_request(next_url, params_copy)
                except Exception:
                    use_collab = False
                    params_copy["fields"] = ",".join(safe_fields)
                    result = _meta_request(next_url, params_copy)
            else:
                params_copy["fields"] = ",".join(safe_fields)
                result = _meta_request(next_url, params_copy)

            for item in result.get("data", []):
                actions         = item.get("actions", [])
                catalog_actions = item.get("catalog_segment_actions", [])
                catalog_values  = item.get("catalog_segment_value", [])

                link_clicks = _extract_action_total(actions, _link_click)
                add_to_cart = _extract_action_total(actions, _add_to_cart)
                follows     = _extract_action_total(actions, _follow)
                engagement  = _extract_action_total(actions, _engagement)
                video_views = _extract_action_total(actions, _video_view)
                initiate_co = _extract_action_total(actions, _initiate_checkout)

                # 구매·매출: 중복 방지 위해 우선순위 첫 번째만
                purchase = _extract_action_first(catalog_actions, _purchase_priority)
                if purchase == 0:
                    purchase = _extract_action_first_fuzzy(catalog_actions,
                        [["shared","purchase"],["purchase","shared"],["purchase"]])

                revenue = _extract_action_first(catalog_values, _purchase_priority)
                if revenue == 0:
                    revenue = _extract_action_first_fuzzy(catalog_values,
                        [["shared","purchase"],["purchase","shared"],["purchase"]])

                rows.append({
                    "날짜":       item.get("date_start",""),
                    "캠페인명":   item.get("campaign_name",""),
                    "광고그룹명": item.get("adset_name",""),
                    "광고명":     item.get("ad_name",""),
                    "비용":       _safe_float(item.get("spend",0)),
                    "노출":       _safe_int(item.get("impressions",0)),
                    "클릭":       _safe_int(link_clicks),
                    "구매":       purchase,
                    "매출액":     revenue,
                    "장바구니":   add_to_cart,
                    "도달":       _safe_int(item.get("reach",0)),
                    "참여":       engagement,
                    "팔로우":     follows,
                    "동영상 조회": video_views,
                    "매체":       "Meta",
                    "_결제시작수": initiate_co,
                })

            next_url   = result.get("paging",{}).get("next")
            next_params = dict(base_params)

    except Exception as e:
        st.error(f"Meta API 오류: {e}")
        return pd.DataFrame()

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.strftime("%Y-%m-%d")
    for col in ["비용","노출","클릭","구매","매출액","장바구니","도달","참여","팔로우","동영상 조회"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df

# ══════════════════════════════════════════════════════════════
# ① 코멘트 생성기 전용 함수
# ══════════════════════════════════════════════════════════════

def fmt_won(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return "0원"
    v = int(round(v))
    if abs(v) >= 10_000: return f"약 {v/10_000:.0f}만 원"
    return f"{v:,}원"

def fmt_roas(spend, revenue):
    if not spend: return "N/A"
    return f"{revenue/spend*100:.0f}%"

def load_report_raw(file):
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    ws = wb["Total_Raw"]
    rows, headers = [], None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if row[0] is None:
            continue
        rows.append(row)
    wb.close()
    df = pd.DataFrame(rows, columns=headers)
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
    for c in ["비용", "구매", "매출액", "클릭", "노출"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if "랜딩페이지" in df.columns:
        df["랜딩페이지"] = df["랜딩페이지"].fillna("미분류").astype(str).str.strip()
    else:
        df["랜딩페이지"] = "미분류"
    return df

def filter_media(df, media_key):
    if media_key == "전체": return df
    kws = COMMENT_MEDIA_MAP[media_key]["campaigns"]
    return df[df["캠페인명"].str.contains("|".join(kws), case=False, na=False)]

def agg_kpi(df):
    return {"spend": df["비용"].sum(), "purchase": df["구매"].sum(),
            "revenue": df["매출액"].sum(), "click": df["클릭"].sum(),
            "impression": df["노출"].sum()}

def daily_agg(df, dt):
    return agg_kpi(df[df["날짜"].dt.date == dt.date()])

def period_agg(df, s, e):
    return agg_kpi(df[(df["날짜"].dt.date >= s.date()) & (df["날짜"].dt.date <= e.date())])

def prev_week_avg(df, dt):
    ws = dt - timedelta(days=dt.weekday() + 7)
    we = ws + timedelta(days=6)
    d = df[(df["날짜"].dt.date >= ws.date()) & (df["날짜"].dt.date <= we.date())]
    if len(d) == 0: return None
    days = max(d["날짜"].dt.date.nunique(), 1)
    sp = d["비용"].sum()
    return {"spend": sp/days, "purchase": d["구매"].sum()/days,
            "revenue": d["매출액"].sum()/days,
            "roas": d["매출액"].sum()/sp*100 if sp > 0 else 0}

def build_topline(label, target_dt, today, prev_day, pw, monthly, weekly, rtype):
    today_roas = fmt_roas(today["spend"], today["revenue"])
    prev_note = ""
    if prev_day and prev_day["spend"] > 0 and today["spend"] > 0:
        diff = today["revenue"]/today["spend"]*100 - prev_day["revenue"]/prev_day["spend"]*100
        if abs(diff) > 5: prev_note = f" (전일 대비 ROAS {diff:+.0f}%p)"
    week_note = ""
    if pw and pw["spend"] > 0 and today["spend"] > 0:
        diff = today["revenue"]/today["spend"]*100 - pw["roas"]
        if abs(diff) > 5: week_note = f", 전주 평균 대비 ROAS {diff:+.0f}%p"
    date_str = (target_dt.strftime("%-m/%-d(%a)")
                .replace("Mon","월").replace("Tue","화").replace("Wed","수")
                .replace("Thu","목").replace("Fri","금").replace("Sat","토").replace("Sun","일"))
    out = f"* {label}\n"
    if rtype in ["daily", "both"]:
        out += (f"- {date_str} 광고비 {fmt_won(today['spend'])} 소진, 구매 건수 {int(today['purchase']):,}건 및 "
                f"매출 {fmt_won(today['revenue'])} 확보 (ROAS {today_roas}){prev_note}{week_note}\n")
    if rtype in ["weekly", "both"] and weekly:
        out += (f"- 주간 광고비 {fmt_won(weekly['spend'])} 소진, 구매 건수 {int(weekly['purchase']):,}건 및 "
                f"매출 {fmt_won(weekly['revenue'])} 확보 (ROAS {fmt_roas(weekly['spend'], weekly['revenue'])})\n")
    if monthly:
        out += (f"- {target_dt.month}월 누적 광고비 {fmt_won(monthly['spend'])} 소진, 구매 건수 {int(monthly['purchase']):,}건 및 "
                f"매출 {fmt_won(monthly['revenue'])} 확보 (ROAS {fmt_roas(monthly['spend'], monthly['revenue'])})\n")
    return out


def _get_few_shot_examples(label, n=5):
    """히스토리에서 동일 매체 최근 코멘트 n개를 few-shot 예시로 반환"""
    history = st.session_state.get("comment_history", [])
    if not history:
        return ""
    same   = [h for h in history if h.get("media","") == label]
    others = [h for h in history if h.get("media","") != label]
    candidates = (same or others)[-n:]
    if not candidates:
        return ""
    lines = ["[과거 코멘트 예시 — 이 형식과 어투를 그대로 따르세요]"]
    for i, h in enumerate(candidates, 1):
        landing = h.get("landing","")
        meta = f"{h.get('date','')} / {h.get('media','')}" + (f"/{landing}" if landing else "")
        lines.append(f"\n--- 예시 {i} ({meta}) ---")
        lines.append(h.get("comment","").strip())
    return "\n".join(lines)


def gen_ai_insight(api_key, today, prev_day, pw, label, target_dt, note=""):
    if not OPENAI_AVAILABLE: return "(openai 패키지 미설치)"
    try:
        client = openai.OpenAI(api_key=api_key)
        roas_t = today["revenue"]/today["spend"]*100 if today["spend"] else 0
        roas_p = prev_day["revenue"]/prev_day["spend"]*100 if prev_day and prev_day["spend"] else 0
        roas_w = pw["roas"] if pw else 0

        few_shot = _get_few_shot_examples(label, n=5)

        system_msg = """당신은 BOJ(뷰티오브조선) 디지털 광고 성과 분석 전문가입니다.
과거 코멘트 예시가 제공되면 그 형식·어투·표현 방식을 정확히 따라서 작성하세요.
예시가 없으면 아래 규칙을 따르세요:
- 2~4줄 bullet(ㄴ 또는 - 시작)
- 유의미한 변화만 언급 (ROAS 5%p 이상 변화)
- "약 X만 원" / "X%p 상승/하락" 형식
- 인사말·서론 없이 bullet로 바로 시작"""

        user_msg = f"""{few_shot}

[분석 대상: {label} / {target_dt.strftime('%m/%d')}]
[오늘] 광고비:{today['spend']:,.0f}원 / 구매:{today['purchase']:.0f}건 / 매출:{today['revenue']:,.0f}원 / ROAS:{roas_t:.0f}%
[전일] 광고비:{prev_day['spend']:,.0f}원 / ROAS:{roas_p:.0f}% / 구매:{prev_day['purchase']:.0f}건
[전주평균] 광고비:{pw['spend']:,.0f}원 / ROAS:{roas_w:.0f}% / 구매:{pw['purchase']:.0f}건
{f'[메모] {note}' if note else ''}

위 데이터로 특이사항 코멘트를 작성하세요."""

        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user",   "content": user_msg},
            ],
            max_tokens=500, temperature=0.3)
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return f"(AI 오류: {str(e)[:80]})"


def render_media_comment(mk, df_m, target_dt, prev_day_dt, month_start, week_start,
                          rtype, api_key, note, all_comments):
    label = COMMENT_MEDIA_MAP[mk]["label"]
    landings_in_media = sorted(df_m["랜딩페이지"].dropna().unique().tolist())
    landings_in_media = [l for l in landings_in_media if l not in ("미분류", "nan", "")]
    has_multi_landing = len(landings_in_media) > 1

    def _render_block(df_sub, sub_label):
        today_d = daily_agg(df_sub, target_dt)
        if today_d["spend"] == 0 and today_d["purchase"] == 0 and today_d["revenue"] == 0:
            return None, None
        prev_d  = daily_agg(df_sub, prev_day_dt)
        pw      = prev_week_avg(df_sub, target_dt)
        monthly = period_agg(df_sub, month_start, target_dt)
        weekly  = period_agg(df_sub, week_start, target_dt) if rtype in ["weekly", "both"] else None
        roas_val = today_d["revenue"]/today_d["spend"]*100 if today_d["spend"] else 0
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("광고비",   fmt_won(today_d["spend"]))
        c2.metric("구매건수", f"{int(today_d['purchase']):,}건")
        c3.metric("매출",     fmt_won(today_d["revenue"]))
        c4.metric("ROAS",     f"{roas_val:.0f}%",
                  delta=f"{roas_val-(pw['roas'] if pw else 0):+.0f}%p vs 전주" if pw else None)
        topline = build_topline(sub_label, target_dt, today_d, prev_d, pw, monthly, weekly, rtype)
        insight = ""
        if api_key and today_d["spend"] > 0:
            with st.spinner(f"{sub_label} AI 분석 중..."):
                insight = gen_ai_insight(
                    api_key, today_d,
                    prev_d if prev_d["spend"] > 0 else {"spend":0,"purchase":0,"revenue":0},
                    pw if pw else {"spend":0,"purchase":0,"revenue":0,"roas":0},
                    sub_label, target_dt, note)
        full = topline + (insight + "\n" if insight else "")
        return full, f"[{sub_label}]\n{full}"

    st.markdown(f"### {label}")
    full, entry = _render_block(df_m, f"{label} (전체)")
    if full is None:
        m_dates = sorted(df_m["날짜"].dt.date.dropna().unique())
        st.warning(f"⚠️ {mk}: {target_dt.date()} 데이터 없음 (최신: {max(m_dates) if m_dates else '없음'})")
        st.divider()
        return
    st.text_area("전체 합산", value=full, height=160, key=f"comment_{mk}_all")
    if entry: all_comments.append(entry)
    if has_multi_landing:
        st.markdown('<div class="landing-header">랜딩별 상세</div>', unsafe_allow_html=True)
        for landing in landings_in_media:
            df_land = df_m[df_m["랜딩페이지"] == landing]
            if df_land.empty: continue
            with st.expander(f"📍 {landing}", expanded=True):
                full_l, entry_l = _render_block(df_land, f"{label} — {landing}")
                if full_l:
                    st.text_area("코멘트", value=full_l, height=160, key=f"comment_{mk}_{landing}")
                    if entry_l: all_comments.append(entry_l)
                else:
                    st.caption("해당 날짜 데이터 없음")
    st.divider()


# ══════════════════════════════════════════════════════════════
# ② 대시보드 전용 함수
# ══════════════════════════════════════════════════════════════

def fmt_krw(v):
    if v is None or (isinstance(v, float) and pd.isna(v)) or v == 0: return "-"
    v = int(round(v))
    if abs(v) >= 100_000_000: return f"{v/100_000_000:.1f}억"
    if abs(v) >= 10_000:      return f"{v/10_000:.0f}만"
    return f"{v:,}"

def fmt_pct(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return "-"
    return f"{v:.1f}%"

def fmt_num(v, dec=0):
    if v is None or (isinstance(v, float) and pd.isna(v)): return "-"
    return f"{v:,.{dec}f}"

def safe_div(a, b, mult=1):
    try:
        r = a/b*mult
        return r if not pd.isna(r) else None
    except: return None

def detect_col(df_cols, candidates):
    for c in candidates:
        if c and c in df_cols: return c
    return None

def normalize_df(df_raw, media):
    """RAW → 대시보드 표준 컬럼 변환 (transform_map 우선, 없으면 컬럼 매핑)"""
    # ① transform 함수가 등록된 매체는 전용 변환 로직 사용
    if media in TRANSFORM_MAP:
        df = TRANSFORM_MAP[media](df_raw).copy()
        # 표준 컬럼명으로 rename (REPORT_COLS → STD_COLS 표시명)
        # 날짜 타입 보장
        df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
        for col in ["비용", "노출", "클릭", "구매", "매출액", "장바구니"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            else:
                df[col] = 0
        # 대시보드에서 쓰는 광고명/소재명 컬럼 추가
        df["광고명/소재명"] = df["광고명"] if "광고명" in df.columns else ""
        # 랜딩페이지 — ADVoost는 네이버 브랜드스토어 단일 랜딩
        df["랜딩페이지"] = "네이버 브랜드스토어"
        df["매체"] = media
        missing = []   # transform 함수가 완전 처리하므로 미매핑 없음
        return df, missing

    # ② 일반 컬럼 매핑 방식
    mapping = MEDIA_COL_MAP.get(media, {})
    cols = df_raw.columns.tolist()
    result, missing = {}, []
    for std_key, std_name in STD_COLS.items():
        found = detect_col(cols, mapping.get(std_key, []))
        if found:
            result[std_name] = df_raw[found]
        else:
            result[std_name] = pd.Series([None]*len(df_raw))
            if std_key not in ["cart", "ad", "landing"]:
                missing.append(std_name)
    df = pd.DataFrame(result)
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce")
    for col in ["비용", "노출", "클릭", "구매", "매출액", "장바구니"]:
        df[col] = pd.to_numeric(
            df[col].astype(str).str.replace(",", "").str.replace("₩", "").str.strip(),
            errors="coerce").fillna(0)
    df["랜딩페이지"] = df["랜딩페이지"].fillna("미분류").astype(str).str.strip()
    df["매체"] = media
    return df, missing

def calc_kpis(df):
    sp=df["비용"].sum(); im=df["노출"].sum(); cl=df["클릭"].sum()
    pu=df["구매"].sum(); rv=df["매출액"].sum()
    return {"spend":sp,"imp":im,"click":cl,"pur":pu,"rev":rv,
            "ctr":safe_div(cl,im,100),"cpc":safe_div(sp,cl),
            "cpm":safe_div(sp,im,1000),"cpa":safe_div(sp,pu),
            "cvr":safe_div(pu,cl,100),"aov":safe_div(rv,pu),"roas":safe_div(rv,sp,100)}

def kpi_table(df_grp):
    rows = []
    for _, row in df_grp.iterrows():
        sp=row.get("비용",0) or 0; im=row.get("노출",0) or 0
        cl=row.get("클릭",0) or 0; pu=row.get("구매",0) or 0; rv=row.get("매출액",0) or 0
        r = {k: row[k] for k in df_grp.columns if k not in ["비용","노출","클릭","구매","매출액","장바구니"]}
        r.update({"비용":int(sp),"노출":int(im),"클릭":int(cl),
                  "CTR":f"{cl/im*100:.2f}%" if im else "-",
                  "CPC":f"{sp/cl:,.0f}" if cl else "-",
                  "CPM":f"{sp/im*1000:,.0f}" if im else "-",
                  "구매":int(pu),
                  "CPA":f"{sp/pu:,.0f}" if pu else "-",
                  "CVR":f"{pu/cl*100:.2f}%" if cl else "-",
                  "매출액":int(rv),
                  "AOV":f"{rv/pu:,.0f}" if pu else "-",
                  "ROAS":f"{rv/sp*100:.0f}%" if sp else "-"})
        rows.append(r)
    return pd.DataFrame(rows)

def section(title):
    st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)

def kpi_row(kpis):
    items = [("광고비",fmt_krw(kpis["spend"])+"만"),("노출",fmt_num(kpis["imp"])),
             ("클릭",fmt_num(kpis["click"])),("구매",fmt_num(kpis["pur"])),
             ("ROAS",fmt_pct(kpis["roas"])),("매출",fmt_krw(kpis["rev"])+"만")]
    for col,(label,val) in zip(st.columns(6), items):
        col.metric(label, val)

def show_table(df, height=320):
    st.dataframe(df, use_container_width=True, height=height, hide_index=True)

def make_agg_table(df, group_cols):
    grp = df.groupby(group_cols).agg(
        비용=("비용","sum"), 노출=("노출","sum"), 클릭=("클릭","sum"),
        구매=("구매","sum"), 매출액=("매출액","sum")
    ).reset_index()
    return kpi_table(grp)


# ══════════════════════════════════════════════════════════════
# 사이드바 네비게이션
# ══════════════════════════════════════════════════════════════
with st.sidebar:
    st.title("📊 BOJ 광고 대시보드")
    page = st.radio("페이지",
                    ["📝 코멘트 생성기", "📊 Sales 리포트",
                     "📈 리포트 대시보드", "🔄 RAW 리포트 변환",
                     "🧠 코멘트 히스토리"],
                    label_visibility="collapsed")
    st.divider()


# ══════════════════════════════════════════════════════════════
# PAGE 1 — 코멘트 생성기
# ══════════════════════════════════════════════════════════════
if page == "📝 코멘트 생성기":

    with st.sidebar:
        st.subheader("⚙️ 설정")
        api_key = st.text_input("OpenAI API Key", type="password", help="없으면 탑라인 멘트만 생성됩니다.")
        st.divider()
        st.subheader("📅 날짜")
        target_date = st.date_input("리포트 날짜", value=datetime.today()-timedelta(days=1))
        st.divider()
        st.subheader("📺 매체")
        selected_media = st.multiselect(
            "코멘트 생성할 매체", options=list(COMMENT_MEDIA_MAP.keys()),
            default=["Meta","Naver BSA","Naver SSA","Naver ADVoost","TikTok"])
        report_type = st.radio("리포트 유형", ["일간","주간+누적","일간+주간+누적"], index=0)
        rtype = {"일간":"daily","주간+누적":"weekly","일간+주간+누적":"both"}[report_type]

    st.title("📝 코멘트 생성기")
    st.caption("최종 리포트 xlsx(Total_Raw 시트 포함)를 업로드하면 탑라인 멘트 + 랜딩별 특이사항을 자동 생성합니다.")

    uploaded = st.file_uploader("📂 최종 리포트 xlsx 업로드", type=["xlsx"])
    if not uploaded:
        st.info("👆 xlsx 파일을 업로드하세요.")
        with st.expander("💡 사용 방법"):
            st.markdown("""
1. 사이드바에서 날짜·매체 설정
2. 최종 리포트 xlsx 업로드 (Total_Raw 시트 필수)
3. **코멘트 생성** 버튼 클릭
4. 각 매체별 **전체 합산** + **랜딩별** 코멘트 확인 후 복사

**랜딩 구분:** Total_Raw의 `랜딩페이지` 컬럼 기준 자동 분리
            """)
        st.stop()

    with st.spinner("데이터 로딩 중..."):
        try:
            df_report = load_report_raw(io.BytesIO(uploaded.read()))
            avail = sorted(df_report["날짜"].dt.date.dropna().unique())
            latest_d, earliest_d = max(avail), min(avail)
            landing_counts = df_report["랜딩페이지"].value_counts()
            landing_info = " | ".join([f"{k}: {v:,}행" for k,v in landing_counts.items()])
            st.success(f"✅ {len(df_report):,}행 | 날짜: {earliest_d} ~ {latest_d}")
            st.caption(f"랜딩 현황: {landing_info}")
        except Exception as e:
            st.error(f"파일 로드 실패: {e}")
            st.stop()

    if target_date > latest_d:
        st.error(f"⚠️ 선택 날짜({target_date}) 데이터 없음. 파일 최신: **{latest_d}**")
        st.stop()
    elif target_date < earliest_d:
        st.warning(f"⚠️ 선택 날짜({target_date})가 시작일({earliest_d})보다 이전입니다.")

    with st.expander("📝 매체별 추가 메모 (AI 힌트용)"):
        notes = {}
        cols2 = st.columns(2)
        for i, m in enumerate(selected_media):
            with cols2[i % 2]:
                notes[m] = st.text_area(m, height=70, key=f"note_{m}",
                                        placeholder="소재명, 이벤트 등 특이사항 힌트")

    if st.button("🚀 코멘트 생성", type="primary", use_container_width=True):
        if not selected_media:
            st.warning("매체를 하나 이상 선택하세요.")
            st.stop()
        target_dt   = datetime.combine(target_date, datetime.min.time())
        prev_day_dt = target_dt - timedelta(days=1)
        month_start = target_dt.replace(day=1)
        week_start  = target_dt - timedelta(days=target_dt.weekday())
        all_comments = []
        prog = st.progress(0)
        for idx, mk in enumerate(selected_media):
            df_m = filter_media(df_report, mk)
            if len(df_m) == 0:
                st.warning(f"{mk}: 데이터 없음")
                prog.progress((idx+1)/len(selected_media))
                continue
            render_media_comment(mk, df_m, target_dt, prev_day_dt,
                                 month_start, week_start, rtype, api_key,
                                 notes.get(mk,""), all_comments)
            prog.progress((idx+1)/len(selected_media))
        if all_comments:
            st.subheader("📋 전체 통합 (복사용)")
            st.text_area("전체", value="\n\n".join(all_comments), height=500, key="all_comments")
        prog.empty()


# ══════════════════════════════════════════════════════════════
# PAGE 2 — Sales 리포트 (이미지 리포트 구조 그대로 구현)
# ══════════════════════════════════════════════════════════════
elif page == "📊 Sales 리포트":

    import json

    # ── 주차 라벨 헬퍼 ────────────────────────────────────────
    def week_label(dt):
        m = dt.month
        # 해당 월의 첫째 날
        first = dt.replace(day=1)
        week_num = (dt.day + first.weekday()) // 7 + 1
        return f"{m}월{week_num}주차"

    # ── 숫자 포맷 헬퍼 ────────────────────────────────────────
    def f_int(v):
        try: return f"{int(round(float(v))):,}"
        except: return "-"
    def f_pct(v, dec=2):
        try: return f"{float(v):.{dec}f}%"
        except: return "-"
    def f_won(v):
        try:
            v = int(round(float(v)))
            if abs(v) >= 10_000: return f"{v/10_000:.0f}만"
            return f"{v:,}"
        except: return "-"

    # ── KPI 행 계산 ───────────────────────────────────────────
    def calc_row(df):
        sp = df["비용"].sum(); im = df["노출"].sum()
        cl = df["클릭"].sum(); pu = df["구매"].sum(); rv = df["매출액"].sum()
        return {
            "Spend": sp, "Imp": im, "Click": cl,
            "CTR":  cl/im*100  if im else 0,
            "CPC":  sp/cl      if cl else 0,
            "CPM":  sp/im*1000 if im else 0,
            "구매 수": pu,
            "CPA (구매)": sp/pu if pu else 0,
            "CVR (구매)": pu/cl*100 if cl else 0,
            "매출": rv,
            "AOV":  rv/pu if pu else 0,
            "ROAS": rv/sp*100 if sp else 0,
        }

    def fmt_row(row):
        """계산된 KPI dict → 표시용 dict"""
        return {
            "Spend":      f_int(row["Spend"]),
            "Imp":        f_int(row["Imp"]),
            "Click":      f_int(row["Click"]),
            "CTR":        f_pct(row["CTR"]),
            "CPC":        f_int(row["CPC"]),
            "CPM":        f_int(row["CPM"]),
            "구매 수":    f_int(row["구매 수"]),
            "CPA (구매)": f_int(row["CPA (구매)"]),
            "CVR (구매)": f_pct(row["CVR (구매)"]),
            "매출":       f_int(row["매출"]),
            "AOV":        f_int(row["AOV"]),
            "ROAS":       f_pct(row["ROAS"], 0) + "%",
        }

    KPI_COLS = ["Spend","Imp","Click","CTR","CPC","CPM","구매 수","CPA (구매)","CVR (구매)","매출","AOV","ROAS"]

    def make_display_df(rows_dict):
        """
        rows_dict: {row_label: raw_kpi_dict}
        반환: 표시용 DataFrame (Campaign 컬럼 + KPI 컬럼)
        """
        records = []
        for label, kpi in rows_dict.items():
            r = {"Campaign": label}
            r.update(fmt_row(kpi))
            records.append(r)
        return pd.DataFrame(records)

    def highlight_total(df):
        """Total 행 강조 스타일"""
        def style_row(row):
            if row.name == len(df)-1 and df.iloc[-1]["Campaign"] in ("Total","total"):
                return ["font-weight:bold; background:#1F4E79; color:white"] * len(row)
            return [""] * len(row)
        return df.style.apply(style_row, axis=1)

    # ── 사이드바: 데이터 소스 선택 & Plan 입력 ────────────────
    with st.sidebar:
        st.subheader("📊 데이터 소스")
        data_src = st.radio("데이터 소스", ["🔄 RAW 변환 데이터", "📂 통합 리포트 xlsx"],
                            key="sales_src", label_visibility="collapsed")

        if data_src == "📂 통합 리포트 xlsx":
            xlsx_up = st.file_uploader("통합 리포트 xlsx", type=["xlsx"], key="sales_xlsx")
        st.divider()

        st.subheader("✏️ Plan / E.Result 입력")
        st.caption("수기 입력 데이터 (RAW에 없는 계획·예상값)")

        plan_spend  = st.number_input("Plan Spend (원)", value=0, step=1000000, format="%d", key="plan_sp")
        plan_pur    = st.number_input("Plan 구매 수",    value=0, step=100,     format="%d", key="plan_pu")
        plan_rev    = st.number_input("Plan 매출 (원)",  value=0, step=1000000, format="%d", key="plan_rv")
        plan_imp    = st.number_input("Plan Imp",        value=0, step=100000,  format="%d", key="plan_im")
        plan_click  = st.number_input("Plan Click",      value=0, step=1000,    format="%d", key="plan_cl")
        er_spend    = st.number_input("E.Result Spend",  value=0, step=1000000, format="%d", key="er_sp")
        er_pur      = st.number_input("E.Result 구매",   value=0, step=100,     format="%d", key="er_pu")
        er_rev      = st.number_input("E.Result 매출",   value=0, step=1000000, format="%d", key="er_rv")
        er_imp      = st.number_input("E.Result Imp",    value=0, step=100000,  format="%d", key="er_im")
        er_click    = st.number_input("E.Result Click",  value=0, step=1000,    format="%d", key="er_cl")

        st.divider()
        st.subheader("📺 매체별 MTD 예산")
        st.caption("예산 수기 입력 후 아래 표에 반영됩니다.")
        mtd_budgets = {}
        mtd_media_list = ["Buzzvil","Criteo","Kakao Biz-Board","Kakao Catalog",
                          "Meta","Naver ADVoost","Naver BSA","Naver SSA","TikTok"]
        for m in mtd_media_list:
            mtd_budgets[m] = st.number_input(f"{m} 예산", value=0, step=100000,
                                              format="%d", key=f"mtd_{m}")

    # ── 데이터 로드 ──────────────────────────────────────────
    df_all = None

    if data_src == "🔄 RAW 변환 데이터":
        if st.session_state.converted_reports:
            dfs = []
            for m, df_c in st.session_state.converted_reports.items():
                df_c2 = df_c.copy(); df_c2["매체"] = m
                dfs.append(df_c2)
            df_all = pd.concat(dfs, ignore_index=True)
            df_all["날짜"] = pd.to_datetime(df_all["날짜"], errors="coerce")
            for c in ["비용","노출","클릭","구매","매출액"]:
                df_all[c] = pd.to_numeric(df_all[c], errors="coerce").fillna(0)
        else:
            st.info("👈 먼저 '🔄 RAW 리포트 변환' 페이지에서 매체별 RAW 파일을 업로드하세요.")
            st.stop()
    else:
        if "xlsx_up" not in dir() or xlsx_up is None:
            st.info("👈 통합 리포트 xlsx를 업로드하세요.")
            st.stop()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_up.read()), data_only=True, read_only=True)
            ws = wb["Total_Raw"]
            rows_raw, headers = [], None
            for row in ws.iter_rows(values_only=True):
                if headers is None: headers = list(row); continue
                if row[0] is None: continue
                rows_raw.append(row)
            wb.close()
            df_all = pd.DataFrame(rows_raw, columns=headers)
            df_all["날짜"] = pd.to_datetime(df_all["날짜"], errors="coerce")
            for c in ["비용","노출","클릭","구매","매출액"]:
                df_all[c] = pd.to_numeric(df_all[c], errors="coerce").fillna(0)
            if "매체" not in df_all.columns and "상품(지면)" in df_all.columns:
                df_all["매체"] = df_all["상품(지면)"]
        except Exception as e:
            st.error(f"파일 로드 실패: {e}")
            st.stop()

    if df_all is None or df_all.empty:
        st.warning("데이터가 없습니다.")
        st.stop()

    # ── 기간 필터 ─────────────────────────────────────────────
    st.title("📊 Sales Campaign Report")
    vd = df_all["날짜"].dropna()
    min_d, max_d = vd.min().date(), vd.max().date()

    fc1, fc2, _ = st.columns([2,2,4])
    with fc1: sel_s = st.date_input("시작일", min_d, min_value=min_d, max_value=max_d, key="rp_s")
    with fc2: sel_e = st.date_input("종료일", max_d, min_value=min_d, max_value=max_d, key="rp_e")

    df_f = df_all[(df_all["날짜"].dt.date >= sel_s) & (df_all["날짜"].dt.date <= sel_e)].copy()

    # ── 랜딩 구분 ─────────────────────────────────────────────
    # converted_reports 기반이면 랜딩페이지 컬럼 없을 수 있음 → 매체로 추정
    NAVER_LANDING   = ["네이버 브랜드스토어"]
    OLIVEYOUNG_LANDING = ["올리브영"]
    KAKAO_LANDING   = ["카카오톡 스토어", "카카오톡 채널", "카카오톡 선물하기"]

    if "랜딩페이지" in df_f.columns:
        df_naver = df_f[df_f["랜딩페이지"].isin(NAVER_LANDING)]
        df_olive = df_f[df_f["랜딩페이지"].isin(OLIVEYOUNG_LANDING)]
        df_kakao = df_f[df_f["랜딩페이지"].isin(KAKAO_LANDING)]
    else:
        # 매체명으로 랜딩 추정
        naver_media = ["Naver ADVoost","Naver BSA","Naver SSA"]
        kakao_media = ["Kakao","Kakao Biz-Board","Kakao Catalog","Kakao Display"]
        df_naver = df_f[df_f["매체"].isin(naver_media)] if "매체" in df_f.columns else df_f.iloc[0:0]
        df_kakao = df_f[df_f["매체"].isin(kakao_media)] if "매체" in df_f.columns else df_f.iloc[0:0]
        df_olive = df_f[~df_f["매체"].isin(naver_media + kakao_media)] if "매체" in df_f.columns else df_f

    # ════════════════════════════════════════════════════════
    # 1. Sales Campaign Overview — 월별
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### ▣ Sales Campaign Overview")

    df_f["월"] = df_f["날짜"].dt.to_period("M").astype(str)
    months = sorted(df_f["월"].unique())

    month_rows = {}
    for m in months:
        month_rows[m] = calc_row(df_f[df_f["월"] == m])
    month_rows["Total"] = calc_row(df_f)

    df_sales_ov = make_display_df(month_rows).rename(columns={"Campaign":"Campaign"})
    st.dataframe(df_sales_ov, use_container_width=True, hide_index=True)

    # ════════════════════════════════════════════════════════
    # 2. Ongoing Campaign Overview — 랜딩별
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### ▣ Ongoing Campaign Overview")

    ongoing_rows = {}
    if not df_naver.empty: ongoing_rows["Naver"] = calc_row(df_naver)
    if not df_olive.empty: ongoing_rows["OliveYoung"] = calc_row(df_olive)
    if not df_kakao.empty: ongoing_rows["Kakao"] = calc_row(df_kakao)
    ongoing_rows["Total"] = calc_row(df_f)

    st.dataframe(make_display_df(ongoing_rows), use_container_width=True, hide_index=True)

    # ════════════════════════════════════════════════════════
    # 3. Plan / E.Result
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### ▣ Plan / E.Result")

    actual = calc_row(df_f)

    # Plan / E.Result 입력값
    plan_kpi = {
        "Spend": plan_spend, "Imp": plan_imp, "Click": plan_click,
        "CTR": plan_click/plan_imp*100 if plan_imp else 0,
        "CPC": plan_spend/plan_click if plan_click else 0,
        "CPM": plan_spend/plan_imp*1000 if plan_imp else 0,
        "구매 수": plan_pur,
        "CPA (구매)": plan_spend/plan_pur if plan_pur else 0,
        "CVR (구매)": plan_pur/plan_click*100 if plan_click else 0,
        "매출": plan_rev,
        "AOV": plan_rev/plan_pur if plan_pur else 0,
        "ROAS": plan_rev/plan_spend*100 if plan_spend else 0,
    }
    er_kpi = {
        "Spend": er_spend, "Imp": er_imp, "Click": er_click,
        "CTR": er_click/er_imp*100 if er_imp else 0,
        "CPC": er_spend/er_click if er_click else 0,
        "CPM": er_spend/er_imp*1000 if er_imp else 0,
        "구매 수": er_pur,
        "CPA (구매)": er_spend/er_pur if er_pur else 0,
        "CVR (구매)": er_pur/er_click*100 if er_click else 0,
        "매출": er_rev,
        "AOV": er_rev/er_pur if er_pur else 0,
        "ROAS": er_rev/er_spend*100 if er_spend else 0,
    }
    # Comparison (실적/plan - 1)
    def comp_val(a, p):
        try:
            r = (float(a) / float(p) - 1) * 100
            return f"{r:+.0f}%"
        except: return "-"

    plan_row = {"Plan / E.Result": "Plan"}; plan_row.update(fmt_row(plan_kpi))
    er_row   = {"Plan / E.Result": "E.Result"}; er_row.update(fmt_row(er_kpi))

    # Comparison row (actual vs plan)
    comp_row = {"Plan / E.Result": "Comparison"}
    for k in KPI_COLS:
        comp_row[k] = comp_val(actual.get(k.replace(" 수","").replace("(구매)","").strip(),0)
                                if k not in ["구매 수","CPA (구매)","CVR (구매)"] else actual.get(k,0),
                                plan_kpi.get(k,0))

    df_plan = pd.DataFrame([plan_row, er_row, comp_row])
    st.dataframe(df_plan, use_container_width=True, hide_index=True)
    if plan_spend == 0 and er_spend == 0:
        st.caption("💡 사이드바에서 Plan / E.Result 수치를 입력하면 자동 계산됩니다.")

    # ════════════════════════════════════════════════════════
    # 4. 매체별 MTD & 예상 마감 효율
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### ▣ 매체 별 MTD & 예상 마감 효율")

    mtd_rows = []
    if "매체" in df_f.columns:
        for m in mtd_media_list:
            df_m = df_f[df_f["매체"] == m]
            k = calc_row(df_m) if not df_m.empty else {kk:0 for kk in ["Spend","Imp","Click","구매 수","CPA (구매)","CVR (구매)","매출","AOV","ROAS","CTR","CPC","CPM"]}
            r = {"Media": m}
            r.update(fmt_row(k))
            r["예산"] = f_int(mtd_budgets.get(m, 0))
            mtd_rows.append(r)

        # E.Result 합계 행
        k_total = calc_row(df_f)
        total_r = {"Media": "E.Result"}
        total_r.update(fmt_row(k_total))
        total_r["예산"] = f_int(er_spend)
        mtd_rows.append(total_r)

    if mtd_rows:
        df_mtd = pd.DataFrame(mtd_rows)
        cols_order = ["Media","예산"] + [c for c in df_mtd.columns if c not in ["Media","예산"]]
        st.dataframe(df_mtd[cols_order], use_container_width=True, hide_index=True)
    st.caption("💡 매체별 예산은 사이드바에서 입력하세요.")

    # ════════════════════════════════════════════════════════
    # 5. 매체별 성과 (월별 합산)
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### ▣ 매체 별 성과")

    if "매체" in df_f.columns:
        media_rows = {}
        for m in sorted(df_f["매체"].unique()):
            media_rows[m] = calc_row(df_f[df_f["매체"] == m])
        media_rows["Total"] = calc_row(df_f)
        df_media_perf = make_display_df(media_rows).rename(columns={"Campaign":"Media"})
        st.dataframe(df_media_perf, use_container_width=True, hide_index=True)
    else:
        st.info("매체 컬럼이 없습니다.")

    # ════════════════════════════════════════════════════════
    # 6. 주차별 성과 + 매출/ROAS 차트
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### ▣ 주차 별 성과")

    df_f["주차"] = df_f["날짜"].apply(lambda d: week_label(d) if pd.notna(d) else "")
    df_f["주차정렬"] = df_f["날짜"].dt.to_period("W").astype(str)
    week_rows = {}
    for ws_sort, ws_label in df_f.groupby("주차정렬")["주차"].first().items():
        grp = df_f[df_f["주차정렬"] == ws_sort]
        week_rows[ws_label] = calc_row(grp)
    week_rows["Total"] = calc_row(df_f)
    df_week_perf = make_display_df(week_rows).rename(columns={"Campaign":"Week"})
    st.dataframe(df_week_perf, use_container_width=True, hide_index=True)

    # 매출/ROAS 일자별 차트 (이미지 2~3번 하단 차트)
    st.markdown("##### 매출 & ROAS 추이")
    try:
        df_daily_chart = df_f.groupby(df_f["날짜"].dt.date).agg(
            매출액=("매출액","sum"), 비용=("비용","sum")
        ).reset_index()
        df_daily_chart.columns = ["날짜","매출액","비용"]
        df_daily_chart["ROAS"] = (df_daily_chart["매출액"] / df_daily_chart["비용"] * 100).round(1)
        df_daily_chart["날짜"] = df_daily_chart["날짜"].astype(str)

        import json as _json
        chart_dates  = df_daily_chart["날짜"].tolist()
        chart_rev    = df_daily_chart["매출액"].tolist()
        chart_roas   = df_daily_chart["ROAS"].tolist()

        chart_html = f"""
<div style="width:100%;padding:8px 0">
<canvas id="salesChart" height="80"></canvas>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<script>
const ctx = document.getElementById('salesChart');
new Chart(ctx, {{
    data: {{
        labels: {_json.dumps(chart_dates)},
        datasets: [
            {{
                type: 'bar',
                label: '매출',
                data: {_json.dumps(chart_rev)},
                backgroundColor: 'rgba(80,80,80,0.75)',
                yAxisID: 'y',
                order: 2,
            }},
            {{
                type: 'line',
                label: 'ROAS',
                data: {_json.dumps(chart_roas)},
                borderColor: 'rgba(220,130,130,0.9)',
                backgroundColor: 'transparent',
                borderWidth: 2,
                pointRadius: 2,
                tension: 0.4,
                yAxisID: 'y1',
                order: 1,
            }}
        ]
    }},
    options: {{
        responsive: true,
        interaction: {{ mode: 'index', intersect: false }},
        plugins: {{
            legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }} }} }}
        }},
        scales: {{
            x: {{ ticks: {{ maxRotation: 45, font: {{ size: 10 }} }} }},
            y: {{
                type: 'linear', position: 'left',
                title: {{ display: true, text: '매출 (원)' }},
                ticks: {{ font: {{ size: 10 }},
                    callback: v => v>=10000 ? (v/10000).toFixed(0)+'만' : v }}
            }},
            y1: {{
                type: 'linear', position: 'right',
                title: {{ display: true, text: 'ROAS (%)' }},
                ticks: {{ font: {{ size: 10 }}, callback: v => v+'%' }},
                grid: {{ drawOnChartArea: false }}
            }}
        }}
    }}
}});
</script>
"""
        st.components.v1.html(chart_html, height=360)
    except Exception as e:
        st.warning(f"차트 생성 오류: {e}")

    # ════════════════════════════════════════════════════════
    # 7. 일자별 성과
    # ════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### ▣ 일자 별 성과")

    df_f["날짜str"] = df_f["날짜"].dt.strftime("%Y-%m-%d")

    # 매체 / 랜딩 필터 (이미지의 슬라이서)
    fc_a, fc_b = st.columns(2)
    with fc_a:
        if "매체" in df_f.columns:
            media_opts = ["전체"] + sorted(df_f["매체"].dropna().unique().tolist())
            sel_media_d = st.multiselect("매체 필터", media_opts[1:], default=[], key="day_media")
        else:
            sel_media_d = []
    with fc_b:
        if "랜딩페이지" in df_f.columns:
            landing_opts = sorted(df_f["랜딩페이지"].dropna().unique().tolist())
            sel_landing_d = st.multiselect("랜딩 필터", landing_opts, default=[], key="day_landing")
        else:
            sel_landing_d = []

    df_day = df_f.copy()
    if sel_media_d and "매체" in df_day.columns:
        df_day = df_day[df_day["매체"].isin(sel_media_d)]
    if sel_landing_d and "랜딩페이지" in df_day.columns:
        df_day = df_day[df_day["랜딩페이지"].isin(sel_landing_d)]

    day_rows = {}
    for ds, grp in df_day.groupby("날짜str"):
        day_rows[ds] = calc_row(grp)
    day_rows["Total"] = calc_row(df_day)
    df_day_perf = make_display_df(day_rows).rename(columns={"Campaign":"Date"})
    st.dataframe(df_day_perf, use_container_width=True, height=420, hide_index=True)


# ══════════════════════════════════════════════════════════════
# PAGE 2 — 리포트 대시보드
# ══════════════════════════════════════════════════════════════
elif page == "📈 리포트 대시보드":

    with st.sidebar:
        st.subheader("📂 RAW 데이터 업로드")
        st.caption("매체별 RAW 파일(xlsx/csv)을 업로드하세요.")
        for media in DASHBOARD_MEDIA_LIST:
            with st.expander(media, expanded=False):
                up = st.file_uploader(f"{media} RAW", type=["xlsx","csv"],
                                      key=f"raw_{media}", label_visibility="collapsed")
                if up:
                    try:
                        df_raw = (pd.read_csv(up, encoding="utf-8-sig")
                                  if up.name.endswith(".csv") else pd.read_excel(up, sheet_name=0))
                        df_norm, miss = normalize_df(df_raw, media)
                        st.session_state.media_data[media]    = df_norm
                        st.session_state.media_warnings[media] = miss
                        if miss: st.warning(f"미매핑: {', '.join(miss)}")
                        else:    st.success(f"✅ {len(df_norm):,}행")
                    except Exception as e:
                        st.error(f"오류: {e}")
        st.divider()
        st.caption("**업로드 현황**")
        for media in DASHBOARD_MEDIA_LIST:
            if media in st.session_state.media_data:
                warn = st.session_state.media_warnings.get(media,[])
                st.caption(f"{'🟡' if warn else '🟢'} {media} — {len(st.session_state.media_data[media]):,}행")
            else:
                st.caption(f"⚪ {media} — 미업로드")
        st.divider()
        if st.button("🗑️ 전체 초기화", use_container_width=True, key="dash_reset"):
            st.session_state.media_data = {}
            st.session_state.media_warnings = {}
            st.rerun()

    st.title("📈 리포트 대시보드")
    st.caption("매체별 RAW 파일을 사이드바에 업로드하면 리포트가 생성됩니다.")

    if not st.session_state.media_data:
        st.info("👈 사이드바에서 매체별 RAW 파일을 업로드하세요.")
        with st.expander("📋 매체별 컬럼 매핑 현황"):
            rows = []
            for m,mapping in MEDIA_COL_MAP.items():
                note = "✅ 전용 변환 로직" if m in TRANSFORM_MAP else ""
                for sk,sn in STD_COLS.items():
                    cands = [c for c in mapping.get(sk,[]) if c]
                    rows.append({"매체":m,"표준 컬럼":sn,
                                 "인식 가능한 RAW 컬럼명":" / ".join(cands) if cands else "❌",
                                 "비고":note if sk=="date" else ""})
            st.dataframe(pd.DataFrame(rows), use_container_width=True, height=400, hide_index=True)
        st.stop()

    all_dfs      = list(st.session_state.media_data.values())
    df_all       = pd.concat(all_dfs, ignore_index=True)
    loaded_media = list(st.session_state.media_data.keys())
    tab_labels   = ["📈 Sales Overview"] + [f"📺 {m}" for m in loaded_media] + ["⚙️ 컬럼 매핑"]
    tabs         = st.tabs(tab_labels)

    # ── Tab 0: Sales Overview ─────────────────────────────────
    with tabs[0]:
        st.subheader("Sales Campaign Overview")
        vd = df_all["날짜"].dropna()
        if vd.empty:
            st.warning("날짜 데이터가 없습니다.")
        else:
            min_d, max_d = vd.min().date(), vd.max().date()
            fc1,fc2,fc3,fc4 = st.columns([2,2,2,2])
            with fc1: sel_start = st.date_input("시작일", min_d, min_value=min_d, max_value=max_d, key="s_s")
            with fc2: sel_end   = st.date_input("종료일", max_d, min_value=min_d, max_value=max_d, key="s_e")
            with fc3:
                m_opts = [LANDING_ALL] + sorted(df_all["매체"].dropna().unique().tolist())
                sel_m  = st.selectbox("매체", m_opts, key="ov_media")
            with fc4:
                l_opts = [LANDING_ALL] + sorted([l for l in df_all["랜딩페이지"].dropna().unique()
                                                 if l not in ("미분류","nan","")])
                sel_l  = st.selectbox("랜딩", l_opts, key="ov_landing")

            df_f = df_all[(df_all["날짜"].dt.date >= sel_start) & (df_all["날짜"].dt.date <= sel_end)].copy()
            if sel_m != LANDING_ALL: df_f = df_f[df_f["매체"] == sel_m]
            if sel_l != LANDING_ALL: df_f = df_f[df_f["랜딩페이지"] == sel_l]

            section("▣ 전체 KPI")
            kpi_row(calc_kpis(df_f))
            st.divider()

            section("▣ 월별 성과")
            df_f["월"] = df_f["날짜"].dt.to_period("M").astype(str)
            show_table(make_agg_table(df_f, ["월"]))
            st.divider()

            section("▣ 매체 × 랜딩 — 월별 성과")
            df_f2 = df_f.copy(); df_f2["월"] = df_f2["날짜"].dt.to_period("M").astype(str)
            show_table(make_agg_table(df_f2, ["매체","랜딩페이지","월"]))
            st.divider()

            section("▣ 매체 × 랜딩 — 주차별 성과")
            df_f3 = df_f.copy()
            df_f3["주차"]  = df_f3["날짜"].dt.to_period("W").apply(
                lambda p: f"{p.start_time.strftime('%m/%d')}~{p.end_time.strftime('%m/%d')}")
            df_f3["_sort"] = df_f3["날짜"].dt.to_period("W").astype(str)
            grp_w = df_f3.groupby(["_sort","매체","랜딩페이지","주차"]).agg(
                비용=("비용","sum"), 노출=("노출","sum"), 클릭=("클릭","sum"),
                구매=("구매","sum"), 매출액=("매출액","sum")
            ).reset_index().sort_values("_sort").drop(columns=["_sort"])
            show_table(kpi_table(grp_w))
            st.divider()

            section("▣ 매체 × 랜딩 — 일자별 성과")
            df_f4 = df_f.copy(); df_f4["날짜str"] = df_f4["날짜"].dt.strftime("%Y-%m-%d")
            grp_d = df_f4.groupby(["날짜str","매체","랜딩페이지"]).agg(
                비용=("비용","sum"), 노출=("노출","sum"), 클릭=("클릭","sum"),
                구매=("구매","sum"), 매출액=("매출액","sum")
            ).reset_index().rename(columns={"날짜str":"날짜"})
            show_table(kpi_table(grp_d), height=400)

    # ── Tab 1~N: 매체별 탭 ───────────────────────────────────
    for ti, media in enumerate(loaded_media):
        with tabs[ti+1]:
            df_m  = st.session_state.media_data[media].copy()
            warns = st.session_state.media_warnings.get(media,[])
            st.subheader(f"{media} 성과 리포트")
            if warns: st.warning(f"⚠️ 미매핑 컬럼: {', '.join(warns)}")

            vd2 = df_m["날짜"].dropna()
            if not vd2.empty:
                mn,mx = vd2.min().date(), vd2.max().date()
                mfc1,mfc2,mfc3,_ = st.columns([2,2,2,2])
                with mfc1: ms = st.date_input("시작일", mn, min_value=mn, max_value=mx, key=f"{media}_s")
                with mfc2: me = st.date_input("종료일", mx, min_value=mn, max_value=mx, key=f"{media}_e")
                df_mf = df_m[(df_m["날짜"].dt.date >= ms) & (df_m["날짜"].dt.date <= me)].copy()
                with mfc3:
                    land_opts = [LANDING_ALL] + sorted([l for l in df_mf["랜딩페이지"].unique()
                                                        if l not in ("미분류","nan","")])
                    sel_land = st.selectbox("랜딩 필터", land_opts, key=f"{media}_land")
                df_mf_f = df_mf[df_mf["랜딩페이지"] == sel_land] if sel_land != LANDING_ALL else df_mf
            else:
                df_mf = df_mf_f = df_m.copy()
                sel_land = LANDING_ALL

            kpi_row(calc_kpis(df_mf_f))
            st.divider()
            it = st.tabs(["그룹별","소재별","주차별","일자별"])

            with it[0]:
                section(f"▣ {media} — 그룹별 성과")
                gc = "광고그룹명"
                if gc in df_mf_f.columns and df_mf_f[gc].notna().any():
                    has_land = df_mf_f["랜딩페이지"].nunique() > 1 and sel_land == LANDING_ALL
                    show_table(make_agg_table(df_mf_f, [gc,"랜딩페이지"] if has_land else [gc]))
                else:
                    st.info("광고그룹명 컬럼이 매핑되지 않았습니다.")

            with it[1]:
                section(f"▣ {media} — 소재별 성과")
                ac = "광고명/소재명"
                if ac in df_mf_f.columns and df_mf_f[ac].notna().any():
                    has_land = df_mf_f["랜딩페이지"].nunique() > 1 and sel_land == LANDING_ALL
                    show_table(make_agg_table(df_mf_f, [ac,"랜딩페이지"] if has_land else [ac]), height=400)
                else:
                    st.info("소재명 컬럼이 매핑되지 않았습니다.")

            with it[2]:
                section(f"▣ {media} — 주차별 성과")
                if df_mf_f["날짜"].notna().any():
                    df_tmp = df_mf_f.copy()
                    df_tmp["주차"] = df_tmp["날짜"].dt.to_period("W").apply(
                        lambda p: f"{p.start_time.strftime('%m/%d')}~{p.end_time.strftime('%m/%d')}")
                    df_tmp["_ws"] = df_tmp["날짜"].dt.to_period("W").astype(str)
                    has_land = df_tmp["랜딩페이지"].nunique() > 1 and sel_land == LANDING_ALL
                    wcols = ["_ws","주차","랜딩페이지"] if has_land else ["_ws","주차"]
                    g = df_tmp.groupby(wcols).agg(비용=("비용","sum"), 노출=("노출","sum"),
                        클릭=("클릭","sum"), 구매=("구매","sum"), 매출액=("매출액","sum")
                    ).reset_index().sort_values("_ws").drop(columns=["_ws"])
                    show_table(kpi_table(g))
                else:
                    st.info("날짜 데이터가 없습니다.")

            with it[3]:
                section(f"▣ {media} — 일자별 성과")
                if df_mf_f["날짜"].notna().any():
                    df_tmp = df_mf_f.copy()
                    df_tmp["날짜str"] = df_tmp["날짜"].dt.strftime("%Y-%m-%d")
                    has_land = df_tmp["랜딩페이지"].nunique() > 1 and sel_land == LANDING_ALL
                    dcols = ["날짜str","랜딩페이지"] if has_land else ["날짜str"]
                    g = df_tmp.groupby(dcols).agg(비용=("비용","sum"), 노출=("노출","sum"),
                        클릭=("클릭","sum"), 구매=("구매","sum"), 매출액=("매출액","sum")
                    ).reset_index().rename(columns={"날짜str":"날짜"})
                    show_table(kpi_table(g), height=400)
                else:
                    st.info("날짜 데이터가 없습니다.")

    # ── Tab 마지막: 컬럼 매핑 설정 ──────────────────────────
    with tabs[-1]:
        st.subheader("⚙️ 컬럼 매핑 확인 및 수동 설정")
        if not st.session_state.media_data:
            st.info("먼저 사이드바에서 RAW 파일을 업로드하세요.")
        else:
            tgt    = st.selectbox("수정할 매체", loaded_media, key="map_media")
            df_s   = st.session_state.media_data[tgt]
            w_cols = st.session_state.media_warnings.get(tgt,[])
            if tgt in TRANSFORM_MAP:
                st.success(f"✅ {tgt}는 전용 변환 로직이 적용됩니다. 수동 매핑 불필요.")
            else:
                st.markdown(f"**미매핑 컬럼:** {', '.join(w_cols) if w_cols else '없음 (정상)'}")
            st.dataframe(df_s.head(5), use_container_width=True, hide_index=True)

            if tgt not in TRANSFORM_MAP:
                st.divider()
                re_up = st.file_uploader(f"{tgt} 재업로드", type=["xlsx","csv"], key=f"remap_{tgt}")
                if re_up:
                    df_re = (pd.read_csv(re_up, encoding="utf-8-sig")
                             if re_up.name.endswith(".csv") else pd.read_excel(re_up, sheet_name=0))
                    raw_cols = ["(없음)"] + df_re.columns.tolist()
                    st.write("RAW 컬럼:", df_re.columns.tolist())
                    manual_map = {}
                    cp = st.columns(2)
                    for i,(sk,sn) in enumerate(STD_COLS.items()):
                        with cp[i%2]:
                            sel = st.selectbox(sn, raw_cols, key=f"man_{tgt}_{sk}")
                            if sel != "(없음)": manual_map[sk] = sel
                    if st.button("✅ 매핑 적용", type="primary"):
                        result, miss2 = {}, []
                        for sk,sn in STD_COLS.items():
                            if sk in manual_map:
                                result[sn] = df_re[manual_map[sk]]
                            else:
                                result[sn] = pd.Series([None]*len(df_re))
                                if sk not in ["cart","ad","landing"]: miss2.append(sn)
                        df_new = pd.DataFrame(result)
                        df_new["날짜"] = pd.to_datetime(df_new["날짜"], errors="coerce")
                        for col in ["비용","노출","클릭","구매","매출액","장바구니"]:
                            df_new[col] = pd.to_numeric(
                                df_new[col].astype(str).str.replace(",","").str.replace("₩","").str.strip(),
                                errors="coerce").fillna(0)
                        df_new["랜딩페이지"] = df_new["랜딩페이지"].fillna("미분류").astype(str).str.strip()
                        df_new["매체"] = tgt
                        st.session_state.media_data[tgt]    = df_new
                        st.session_state.media_warnings[tgt] = miss2
                        st.success("✅ 적용 완료!")
                        st.rerun()

        st.divider()
        st.markdown("**전체 컬럼 매핑 현황**")
        mr = []
        for m,mapping in MEDIA_COL_MAP.items():
            for sk,sn in STD_COLS.items():
                cands = [c for c in mapping.get(sk,[]) if c]
                mr.append({"매체":m,"표준 컬럼":sn,
                           "인식 가능한 컬럼명":" / ".join(cands) if cands else "❌",
                           "변환방식":"전용 함수" if (m in TRANSFORM_MAP and sk=="date") else ""})
        st.dataframe(pd.DataFrame(mr), use_container_width=True, height=400, hide_index=True)


# ══════════════════════════════════════════════════════════════
# PAGE 3 — RAW 리포트 변환
# ══════════════════════════════════════════════════════════════
elif page == "🔄 RAW 리포트 변환":

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── xlsx 빌더 ─────────────────────────────────────────────
    def build_xlsx(df_out, sheet_title="Report"):
        HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
        HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        DATA_FONT   = Font(name="Arial", size=10)
        CENTER = Alignment(horizontal="center", vertical="center")
        LEFT   = Alignment(horizontal="left",   vertical="center")
        RIGHT  = Alignment(horizontal="right",  vertical="center")
        THIN   = Side(style="thin", color="D0D0D0")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
        NUM_FMT = {"비용":"#,##0.00","노출":"#,##0","클릭":"#,##0",
                   "구매":"#,##0","매출액":"#,##0","장바구니":"#,##0",
                   "도달":"#,##0","참여":"#,##0","팔로우":"#,##0","동영상 조회":"#,##0"}
        COL_W = {"날짜":13,"캠페인명":22,"광고그룹명":20,"광고명":20,
                 "비용":14,"노출":12,"클릭":10,"구매":10,"매출액":14,
                 "장바구니":12,"도달":10,"참여":10,"팔로우":10,"동영상 조회":13}
        wb = Workbook(); ws = wb.active; ws.title = sheet_title[:31]
        for ci, col in enumerate(REPORT_COLS, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font=HEADER_FONT; cell.fill=HEADER_FILL
            cell.alignment=CENTER; cell.border=BORDER
        for ri, row in df_out.reset_index(drop=True).iterrows():
            fill = ALT_FILL if ri % 2 == 1 else None
            for ci, col in enumerate(REPORT_COLS, 1):
                val = row.get(col)
                if isinstance(val, float) and pd.isna(val): val = None
                cell = ws.cell(row=ri+2, column=ci, value=val)
                cell.font=DATA_FONT; cell.border=BORDER
                if fill: cell.fill=fill
                cell.alignment = RIGHT if col in NUM_FMT else (CENTER if col=="날짜" else LEFT)
                if col in NUM_FMT and val is not None:
                    cell.number_format = NUM_FMT[col]
        for ci, col in enumerate(REPORT_COLS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(col, 12)
        ws.row_dimensions[1].height = 22
        for ri in range(2, len(df_out)+2): ws.row_dimensions[ri].height = 18
        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    def make_kpi_summary(df_c):
        df_k = df_c.copy()
        df_k["날짜str"] = pd.to_datetime(df_k["날짜"], errors="coerce").dt.strftime("%Y-%m-%d")
        grp = df_k.groupby("날짜str").agg(
            비용=("비용","sum"), 노출=("노출","sum"), 클릭=("클릭","sum"),
            구매=("구매","sum"), 매출액=("매출액","sum"), 장바구니=("장바구니","sum")
        ).reset_index().rename(columns={"날짜str":"날짜"})
        rows_k = []
        for _, r in grp.iterrows():
            sp=r["비용"]; cl=r["클릭"]; pu=r["구매"]; rv=r["매출액"]
            rows_k.append({
                "날짜":f"{r['날짜']}","비용":f"{sp:,.0f}",
                "노출":f"{int(r['노출']):,}","클릭":f"{int(cl):,}",
                "CTR":f"{cl/r['노출']*100:.2f}%" if r["노출"] else "-",
                "구매":f"{int(pu):,}","CVR":f"{pu/cl*100:.2f}%" if cl else "-",
                "매출액":f"{int(rv):,}","ROAS":f"{rv/sp*100:.0f}%" if sp else "-",
                "장바구니":f"{int(r['장바구니']):,}",
            })
        return pd.DataFrame(rows_k)

    # ── 사이드바 ─────────────────────────────────────────────
    with st.sidebar:
        st.subheader("🔄 변환 현황")
        if st.session_state.converted_reports:
            for m, df_c in st.session_state.converted_reports.items():
                dates = pd.to_datetime(df_c["날짜"], errors="coerce").dt.date.dropna()
                dr = f"{dates.min()} ~ {dates.max()}" if not dates.empty else "날짜 없음"
                st.caption(f"🟢 {m} — {len(df_c):,}행 ({dr})")
        else:
            st.caption("아직 변환된 데이터 없음")
        st.divider()
        if st.button("🗑️ 변환 데이터 초기화", use_container_width=True, key="conv_reset"):
            st.session_state.converted_reports = {}
            st.rerun()

    # ── 메인 ─────────────────────────────────────────────────
    st.title("🔄 RAW 리포트 변환")
    st.caption("매체별 RAW 파일을 한꺼번에 업로드하면 매체를 자동 감지해서 취합합니다.")

    # 지원 매체 안내
    with st.expander("📋 지원 매체 및 컬럼 매핑 확인", expanded=False):
        st.markdown("**현재 자동 감지 가능한 매체:**")
        sig_rows = []
        for media, sig_cols in MEDIA_SIGNATURES:
            sig_rows.append({
                "매체": media,
                "자동 감지 기준 컬럼": " / ".join(sig_cols),
                "변환 함수": "✅ 등록됨" if media in TRANSFORM_MAP else "❌ 미등록",
            })
        st.dataframe(pd.DataFrame(sig_rows), use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("**매체별 컬럼 매핑:**")
        rule_rows = []
        MEDIA_RULES = {
            "Naver ADVoost": [
                ("기간", "날짜", "다양한 형식 자동 파싱"),
                ("캠페인 이름", "캠페인명", "그대로"),
                ("애셋 그룹 이름", "광고그룹명", "그대로"),
                ("(없음)", "광고명", "Asset_all_items 고정"),
                ("총 비용", "비용", "÷ 1.1 (VAT 제외)"),
                ("노출", "노출", "그대로"),
                ("클릭", "클릭", "그대로"),
                ("구매완료수", "구매", "그대로"),
                ("구매완료 전환 매출액", "매출액", "그대로"),
                ("장바구니 담기수", "장바구니", "그대로"),
                ("—", "도달·참여·팔로우·동영상 조회", "빈 칸"),
            ],
            "Meta": [
                ("일", "날짜", "YYYY-MM-DD 그대로"),
                ("캠페인 이름", "캠페인명", "그대로"),
                ("광고 세트 이름", "광고그룹명", "그대로"),
                ("광고 이름", "광고명", "그대로"),
                ("지출 금액 (KRW)", "비용", "그대로 (VAT 없음)"),
                ("노출", "노출", "그대로"),
                ("링크 클릭", "클릭", "NaN→0"),
                ("공유 항목이 포함된 구매", "구매", "없으면 0"),
                ("공유 항목의 구매 전환값", "매출액", "없으면 0"),
                ("공유 항목이 포함된 장바구니에 담기", "장바구니", "없으면 0"),
                ("도달", "도달", "없으면 0"),
                ("게시물 참여", "참여", "없으면 빈 칸"),
                ("Instagram 팔로우", "팔로우", "없으면 빈 칸"),
                ("동영상 3초 이상 재생", "동영상 조회", "없으면 0"),
            ],
        }
        for media, rules in MEDIA_RULES.items():
            for raw_col, std_col, note in rules:
                rule_rows.append({"매체": media, "RAW 컬럼": raw_col, "표준 컬럼": std_col, "처리": note})
        st.dataframe(pd.DataFrame(rule_rows), use_container_width=True, height=320, hide_index=True)

    # ── Meta API 자동 수집 ───────────────────────────────────
    section("▣ Meta API 자동 수집 (Secrets 설정 시 사용 가능)")

    with st.expander("📡 Meta API로 데이터 가져오기", expanded=False):
        st.caption("Streamlit Secrets에 `META_ACCESS_TOKEN`과 `META_AD_ACCOUNT_ID`가 설정되어 있어야 합니다.")

        # Secrets 설정 여부 확인
        has_secrets = False
        try:
            _ = st.secrets["META_ACCESS_TOKEN"]
            _ = st.secrets["META_AD_ACCOUNT_ID"]
            has_secrets = True
            st.success("✅ Meta API 인증 정보가 설정되어 있습니다.")
        except Exception:
            st.warning("⚠️ Streamlit Secrets에 META_ACCESS_TOKEN / META_AD_ACCOUNT_ID가 없습니다.")
            st.code("""# .streamlit/secrets.toml 예시
META_ACCESS_TOKEN  = "EAAxxxxxxx..."
META_AD_ACCOUNT_ID = "123456789"
""")

        mc1, mc2 = st.columns(2)
        with mc1:
            meta_start = st.date_input("시작일", value=datetime.today()-timedelta(days=7), key="meta_api_start")
        with mc2:
            meta_end   = st.date_input("종료일", value=datetime.today()-timedelta(days=1), key="meta_api_end")

        if st.button("📥 Meta 데이터 가져오기", type="primary",
                     use_container_width=True, key="meta_fetch_btn",
                     disabled=not has_secrets):
            with st.spinner("Meta API에서 데이터를 가져오는 중..."):
                df_meta_fetched = fetch_meta_data(
                    str(meta_start), str(meta_end)
                )
            if not df_meta_fetched.empty:
                st.session_state.meta_api_df = df_meta_fetched

                # converted_reports에 Meta로 병합 (기존 파일 업로드와 동일하게 취급)
                REPORT_COLS_META = ["날짜","캠페인명","광고그룹명","광고명","비용",
                                    "노출","클릭","구매","매출액","장바구니",
                                    "도달","참여","팔로우","동영상 조회"]
                df_meta_std = df_meta_fetched.copy()
                # 표준 컬럼만 유지 (없는 컬럼은 None)
                for col in REPORT_COLS_META:
                    if col not in df_meta_std.columns:
                        df_meta_std[col] = None
                df_meta_std = df_meta_std[REPORT_COLS_META]
                df_meta_std["날짜"] = pd.to_datetime(df_meta_std["날짜"], errors="coerce")

                if "Meta" in st.session_state.converted_reports:
                    existing = st.session_state.converted_reports["Meta"]
                    df_meta_std = pd.concat([existing, df_meta_std], ignore_index=True).drop_duplicates(
                        subset=["날짜","캠페인명","광고그룹명","광고명"]
                    ).sort_values(["날짜","캠페인명","광고그룹명","광고명"]).reset_index(drop=True)

                st.session_state.converted_reports["Meta"] = df_meta_std
                st.success(f"✅ {len(df_meta_fetched):,}행 수집 완료 → converted_reports['Meta']에 저장됨")

                col_p, col_d = st.columns(2)
                with col_p:
                    dates = pd.to_datetime(df_meta_fetched["날짜"], errors="coerce").dt.date.dropna()
                    st.metric("수집 기간", f"{dates.min()} ~ {dates.max()}" if not dates.empty else "-")
                with col_d:
                    st.metric("총 행수", f"{len(df_meta_fetched):,}")

                st.dataframe(df_meta_fetched.head(10), use_container_width=True,
                             height=220, hide_index=True)
            else:
                st.warning("수집된 데이터가 없습니다. 날짜 범위와 Secrets 설정을 확인하세요.")

        # 이미 수집된 데이터가 있으면 현황 표시
        if not st.session_state.meta_api_df.empty:
            df_prev = st.session_state.meta_api_df
            dates = pd.to_datetime(df_prev["날짜"], errors="coerce").dt.date.dropna()
            st.caption(f"현재 수집된 Meta API 데이터: **{len(df_prev):,}행** "
                       f"({dates.min()} ~ {dates.max()})" if not dates.empty else "")
            if st.button("🗑️ Meta API 데이터 초기화", key="meta_api_reset"):
                st.session_state.meta_api_df = pd.DataFrame()
                if "Meta" in st.session_state.converted_reports:
                    del st.session_state.converted_reports["Meta"]
                st.rerun()

    # ── 파일 업로드 (매체 구분 없이 한꺼번에) ─────────────────
    section("▣ RAW 파일 업로드 — 매체 구분 없이 모두 업로드하세요")

    all_files = st.file_uploader(
        "RAW 파일 업로드 (여러 매체, 여러 파일 동시 가능)",
        type=["csv", "xlsx"],
        accept_multiple_files=True,
        key="conv_upload_all",
        label_visibility="collapsed",
    )

    if all_files:
        # ── 파일별 처리 ──────────────────────────────────────
        results_by_media = {}   # {매체명: [df1, df2, ...]}
        errors = []
        detection_log = []

        for f in all_files:
            try:
                df_raw = (pd.read_csv(f, encoding="utf-8-sig")
                          if f.name.endswith(".csv") else pd.read_excel(f, sheet_name=0))
                df_raw.columns = [str(c).strip() for c in df_raw.columns]

                # 매체 자동 감지
                detected = detect_media(df_raw)

                if detected is None:
                    errors.append(f"❌ `{f.name}` — 매체 감지 실패 (알 수 없는 컬럼 구조)")
                    detection_log.append({"파일": f.name, "감지 매체": "❌ 감지 실패", "행수": len(df_raw), "컬럼 수": len(df_raw.columns)})
                    continue

                if detected not in TRANSFORM_MAP:
                    errors.append(f"⚠️ `{f.name}` — {detected} 감지됐으나 변환 함수 미등록")
                    detection_log.append({"파일": f.name, "감지 매체": f"⚠️ {detected} (미등록)", "행수": len(df_raw), "컬럼 수": len(df_raw.columns)})
                    continue

                # 변환
                df_converted = TRANSFORM_MAP[detected](df_raw)
                if detected not in results_by_media:
                    results_by_media[detected] = []
                results_by_media[detected].append(df_converted)
                detection_log.append({"파일": f.name, "감지 매체": f"✅ {detected}", "행수": len(df_converted), "컬럼 수": len(df_raw.columns)})

            except Exception as e:
                errors.append(f"❌ `{f.name}` 오류: {e}")
                detection_log.append({"파일": f.name, "감지 매체": "❌ 오류", "행수": 0, "컬럼 수": 0})

        # 감지 결과 로그 표시
        section("▣ 파일 감지 결과")
        st.dataframe(pd.DataFrame(detection_log), use_container_width=True, hide_index=True)
        for err in errors:
            st.warning(err)

        if not results_by_media:
            st.error("변환 가능한 파일이 없습니다. 지원 매체 컬럼 매핑을 확인하세요.")
            st.stop()

        # ── 매체별 취합 및 session_state 저장 ────────────────
        section("▣ 매체별 취합 결과")
        for media, dfs in results_by_media.items():
            df_merged = pd.concat(dfs, ignore_index=True).sort_values(
                ["날짜", "캠페인명", "광고그룹명", "광고명"]
            ).reset_index(drop=True)

            # 기존 데이터와 병합 (중복 제거)
            if media in st.session_state.converted_reports:
                existing = st.session_state.converted_reports[media]
                df_merged = pd.concat([existing, df_merged], ignore_index=True).drop_duplicates(
                    subset=["날짜", "캠페인명", "광고그룹명", "광고명"]
                ).sort_values(["날짜", "캠페인명", "광고그룹명", "광고명"]).reset_index(drop=True)

            st.session_state.converted_reports[media] = df_merged

            dates = pd.to_datetime(df_merged["날짜"], errors="coerce").dt.date.dropna()
            dr = f"{dates.min()} ~ {dates.max()}" if not dates.empty else "날짜 없음"
            mc1, mc2, mc3 = st.columns([2, 2, 4])
            mc1.metric(media, f"{len(df_merged):,}행")
            mc2.metric("날짜 범위", dr)
            mc3.metric("파일 수", f"{len(dfs)}개")

        # ── 전체 취합 ─────────────────────────────────────────
        all_converted = []
        for m, df_c in st.session_state.converted_reports.items():
            df_c2 = df_c.copy(); df_c2["매체"] = m
            all_converted.append(df_c2)

        df_total = pd.concat(all_converted, ignore_index=True).sort_values(
            ["날짜", "매체", "캠페인명", "광고그룹명", "광고명"]
        ).reset_index(drop=True)

        st.success(f"✅ 전체 취합 완료 — {len(st.session_state.converted_reports)}개 매체 / 총 {len(df_total):,}행")

        # ── 미리보기 ──────────────────────────────────────────
        section("▣ 통합 데이터 미리보기")
        preview_cols = ["날짜", "매체"] + [c for c in REPORT_COLS if c != "날짜"]
        st.dataframe(df_total[preview_cols].head(50), use_container_width=True, height=280, hide_index=True)

        # ── KPI 요약 ──────────────────────────────────────────
        section("▣ 날짜 × 매체별 KPI 요약")
        df_t2 = df_total.copy()
        df_t2["날짜str"] = pd.to_datetime(df_t2["날짜"], errors="coerce").dt.strftime("%Y-%m-%d")
        grp_t = df_t2.groupby(["날짜str","매체"]).agg(
            비용=("비용","sum"), 노출=("노출","sum"), 클릭=("클릭","sum"),
            구매=("구매","sum"), 매출액=("매출액","sum")
        ).reset_index().rename(columns={"날짜str":"날짜"})
        rows_t = []
        for _, r in grp_t.iterrows():
            sp=r["비용"]; cl=r["클릭"]; pu=r["구매"]; rv=r["매출액"]
            rows_t.append({
                "날짜":r["날짜"],"매체":r["매체"],
                "비용":f"{sp:,.0f}","노출":f"{int(r['노출']):,}","클릭":f"{int(cl):,}",
                "CTR":f"{cl/r['노출']*100:.2f}%" if r["노출"] else "-",
                "구매":f"{int(pu):,}","CVR":f"{pu/cl*100:.2f}%" if cl else "-",
                "매출액":f"{int(rv):,}","ROAS":f"{rv/sp*100:.0f}%" if sp else "-",
            })
        st.dataframe(pd.DataFrame(rows_t), use_container_width=True, hide_index=True)

        # ── 다운로드 ──────────────────────────────────────────
        section("▣ 리포트 다운로드")

        def build_integrated_xlsx():
            from openpyxl import Workbook as WB2
            from openpyxl.styles import Font as F2, PatternFill as PF2, Alignment as AL2, Border as BD2, Side as SD2
            from openpyxl.utils import get_column_letter as gcl2

            HF = PF2("solid", fgColor="1F4E79")
            HFnt = F2(name="Arial", bold=True, color="FFFFFF", size=10)
            DF = F2(name="Arial", size=10)
            C = AL2(horizontal="center", vertical="center")
            L = AL2(horizontal="left",   vertical="center")
            R = AL2(horizontal="right",  vertical="center")
            TH = SD2(style="thin", color="D0D0D0")
            BR = BD2(left=TH, right=TH, top=TH, bottom=TH)
            AF = PF2("solid", fgColor="EBF3FB")
            NF = {"비용":"#,##0.00","노출":"#,##0","클릭":"#,##0","구매":"#,##0",
                  "매출액":"#,##0","장바구니":"#,##0","도달":"#,##0",
                  "참여":"#,##0","팔로우":"#,##0","동영상 조회":"#,##0"}
            CW = {"날짜":13,"캠페인명":22,"광고그룹명":20,"광고명":20,"비용":14,
                  "노출":12,"클릭":10,"구매":10,"매출액":14,"장바구니":12,
                  "도달":10,"참여":10,"팔로우":10,"동영상 조회":13}

            def write_ws(ws, df_out):
                for ci, col in enumerate(REPORT_COLS, 1):
                    cell = ws.cell(row=1, column=ci, value=col)
                    cell.font=HFnt; cell.fill=HF; cell.alignment=C; cell.border=BR
                for ri, row in df_out.reset_index(drop=True).iterrows():
                    fill = AF if ri % 2 == 1 else None
                    for ci, col in enumerate(REPORT_COLS, 1):
                        val = row.get(col)
                        if isinstance(val, float) and pd.isna(val): val = None
                        cell = ws.cell(row=ri+2, column=ci, value=val)
                        cell.font=DF; cell.border=BR
                        if fill: cell.fill=fill
                        cell.alignment = R if col in NF else (C if col=="날짜" else L)
                        if col in NF and val is not None: cell.number_format = NF[col]
                for ci, col in enumerate(REPORT_COLS, 1):
                    ws.column_dimensions[gcl2(ci)].width = CW.get(col, 12)
                ws.row_dimensions[1].height = 22
                for ri in range(2, len(df_out)+2): ws.row_dimensions[ri].height = 18
                ws.freeze_panes = "A2"

            wb = WB2(); wb.remove(wb.active)
            # 매체별 시트
            for m, df_m in st.session_state.converted_reports.items():
                write_ws(wb.create_sheet(title=m[:31]), df_m)
            # Total_Raw 시트
            write_ws(wb.create_sheet(title="Total_Raw"), df_total[REPORT_COLS])
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return buf

        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                label="⬇️ 통합 리포트 xlsx (매체별 시트 + Total_Raw)",
                data=build_integrated_xlsx(),
                file_name=f"통합리포트_{datetime.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True,
            )
        with dl2:
            st.download_button(
                label="⬇️ Total_Raw만 xlsx",
                data=build_xlsx(df_total[REPORT_COLS], sheet_title="Total_Raw"),
                file_name=f"Total_Raw_{datetime.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    else:
        st.info("👆 여러 매체의 RAW 파일을 한꺼번에 업로드하세요. 매체를 자동으로 감지합니다.")
        st.markdown("""
**업로드 방법:**
- Meta csv, ADVoost csv, TikTok csv 등을 **한 번에 모두 선택**해서 업로드
- 파일명은 무관 — **컬럼 구조**로 매체를 자동 판별

**현재 지원 매체:** Naver ADVoost, Meta

**추후 추가 예정:** TikTok, Kakao, Naver SSA/BSA, Criteo, Buzzvil
        """)


# ══════════════════════════════════════════════════════════════
# PAGE 4 — 코멘트 히스토리
# ══════════════════════════════════════════════════════════════
elif page == "🧠 코멘트 히스토리":
    import json, re

    st.title("🧠 코멘트 히스토리")
    st.caption("과거 코멘트를 저장해두면 AI가 동일한 형식·어투로 코멘트를 생성합니다.")

    # ── 사이드바: 현황 ────────────────────────────────────────
    with st.sidebar:
        st.subheader("📊 히스토리 현황")
        history = st.session_state.comment_history
        if history:
            df_h = pd.DataFrame(history)
            st.metric("총 저장 건수", f"{len(history)}건")
            if "media" in df_h.columns:
                for m, cnt in df_h["media"].value_counts().items():
                    st.caption(f"  {m}: {cnt}건")
        else:
            st.caption("저장된 히스토리 없음")
        st.divider()
        if st.button("🗑️ 히스토리 전체 삭제", use_container_width=True, key="hist_reset"):
            st.session_state.comment_history = []
            st.rerun()

    # ── 탭 구성 ──────────────────────────────────────────────
    tab_upload, tab_manual, tab_view = st.tabs(
        ["📂 파일 업로드 (txt/csv)", "✏️ 직접 입력·저장", "📋 저장된 히스토리 보기"]
    )

    # ════════════════════════════════════════════════════════
    # TAB 1 — 파일 업로드
    # ════════════════════════════════════════════════════════
    with tab_upload:
        st.subheader("파일로 히스토리 업로드")

        with st.expander("📌 파일 형식 안내", expanded=True):
            st.markdown("""
**지원 형식 1: txt 파일** — 카카오톡·슬랙에서 복사한 텍스트 그대로 붙여넣고 저장  
날짜/매체 구분자가 없어도 됩니다. 아래처럼 블록 단위로 구성하면 자동 파싱됩니다.

```
[2026-03-01 / Meta / 올리브영]
* Meta — 올리브영
- 3/1(일) 광고비 약 237만 원 소진, 구매 건수 284건 ...
ㄴ 전주 대비 CVR 급증 ...

[2026-03-02 / Meta / 네이버 브랜드스토어]
* Meta — 네이버 브랜드스토어
- 3/2(월) 광고비 약 150만 원 소진 ...
```

**지원 형식 2: csv 파일** — 헤더: `date, media, landing, comment`
```
date,media,landing,comment
2026-03-01,Meta,올리브영,"* Meta...\\nㄴ ..."
2026-03-02,Naver BSA,네이버 브랜드스토어,"* Naver BSA..."
```

**형식 3: 구분자 없는 txt** — 날짜·매체를 직접 지정하고 전체 텍스트를 일괄 저장
            """)

        uploaded_hist = st.file_uploader(
            "히스토리 파일 업로드", type=["txt","csv"], key="hist_upload"
        )

        # 날짜·매체 수동 지정 (txt 구분자 없을 때)
        col_m1, col_m2, col_m3 = st.columns(3)
        with col_m1:
            manual_date = st.text_input("날짜 (YYYY-MM-DD)", placeholder="2026-03-01", key="hist_date")
        with col_m2:
            manual_media = st.selectbox("매체", [""] + list(COMMENT_MEDIA_MAP.keys()), key="hist_media")
        with col_m3:
            manual_landing = st.text_input("랜딩 (선택)", placeholder="올리브영", key="hist_landing")

        if uploaded_hist and st.button("📥 파일 파싱 & 저장", type="primary", use_container_width=True, key="hist_parse"):
            try:
                raw_text = uploaded_hist.read().decode("utf-8-sig").strip()
                added = []

                if uploaded_hist.name.endswith(".csv"):
                    # ── CSV 파싱 ──────────────────────────────
                    import csv, io as _io
                    reader = csv.DictReader(_io.StringIO(raw_text))
                    for row in reader:
                        entry = {
                            "date":    row.get("date","").strip(),
                            "media":   row.get("media","").strip(),
                            "landing": row.get("landing","").strip(),
                            "comment": row.get("comment","").strip(),
                        }
                        if entry["comment"]:
                            added.append(entry)

                else:
                    # ── TXT 파싱 ─────────────────────────────
                    # 패턴: [YYYY-MM-DD / 매체 / 랜딩] 으로 블록 분리
                    block_pattern = re.compile(
                        r'\[(\d{4}-\d{2}-\d{2})\s*/\s*([^/\]]+?)(?:\s*/\s*([^\]]*?))?\]',
                        re.MULTILINE
                    )
                    matches = list(block_pattern.finditer(raw_text))

                    if matches:
                        # 구분자 있는 txt
                        for i, m in enumerate(matches):
                            block_start = m.end()
                            block_end   = matches[i+1].start() if i+1 < len(matches) else len(raw_text)
                            comment_text = raw_text[block_start:block_end].strip()
                            if comment_text:
                                added.append({
                                    "date":    m.group(1).strip(),
                                    "media":   m.group(2).strip(),
                                    "landing": (m.group(3) or "").strip(),
                                    "comment": comment_text,
                                })
                    else:
                        # 구분자 없는 txt — 수동 지정값 사용, 전체를 하나의 코멘트로
                        if not manual_date or not manual_media:
                            st.warning("구분자가 없는 txt 파일입니다. 위에서 날짜와 매체를 지정하세요.")
                        else:
                            added.append({
                                "date":    manual_date.strip(),
                                "media":   manual_media.strip(),
                                "landing": manual_landing.strip(),
                                "comment": raw_text,
                            })

                if added:
                    # 중복 제거 후 추가
                    existing_keys = {
                        (h["date"], h["media"], h["landing"], h["comment"][:50])
                        for h in st.session_state.comment_history
                    }
                    new_entries = [
                        e for e in added
                        if (e["date"], e["media"], e["landing"], e["comment"][:50])
                        not in existing_keys
                    ]
                    st.session_state.comment_history.extend(new_entries)
                    st.session_state.comment_history.sort(key=lambda x: x.get("date",""))
                    st.success(f"✅ {len(new_entries)}건 저장 완료 (중복 {len(added)-len(new_entries)}건 제외)")
                    if new_entries:
                        st.dataframe(
                            pd.DataFrame(new_entries)[["date","media","landing","comment"]],
                            use_container_width=True, height=200, hide_index=True
                        )
                else:
                    st.warning("저장할 코멘트를 찾지 못했습니다. 파일 형식을 확인하세요.")

            except Exception as e:
                st.error(f"파싱 오류: {e}")
                import traceback; st.code(traceback.format_exc())

        # 템플릿 다운로드
        st.divider()
        st.caption("📎 아래 템플릿을 받아서 작성 후 업로드하세요.")
        template_txt = """[2026-03-01 / Meta / 올리브영]
* Meta — 올리브영
- 3/1(일) 광고비 약 237만 원 소진, 구매 건수 284건 및 매출 약 539만 원 확보 (ROAS 226%)
ㄴ 전주 대비 세트 공통적으로 CVR 급증, 기존 고효율 소재 위주 구매 획득
ㄴ 특히, 전주 효율 저조했던 잠재고객 세트의 경우 ROAS 2배 이상 개선

[2026-03-02 / Naver BSA / 네이버 브랜드스토어]
* Naver BSA
- 3/2(월) 광고비 약 150만 원 소진, 구매 건수 173건 및 매출 약 856만 원 확보 (ROAS 570%)
ㄴ 전주와 CVR 동수준 유지, AOV 1.3만원 높게 형성되며 매출 상승
"""
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            st.download_button(
                "⬇️ txt 템플릿 다운로드",
                data=template_txt.encode("utf-8-sig"),
                file_name="comment_history_template.txt",
                mime="text/plain", use_container_width=True,
            )
        with col_t2:
            template_csv = "date,media,landing,comment\n2026-03-01,Meta,올리브영,\"* Meta\\nㄴ 예시 코멘트\"\n"
            st.download_button(
                "⬇️ csv 템플릿 다운로드",
                data=template_csv.encode("utf-8-sig"),
                file_name="comment_history_template.csv",
                mime="text/csv", use_container_width=True,
            )

    # ════════════════════════════════════════════════════════
    # TAB 2 — 직접 입력·저장
    # ════════════════════════════════════════════════════════
    with tab_manual:
        st.subheader("코멘트 직접 입력")
        st.caption("코멘트를 복사해서 붙여넣고, 날짜·매체 정보를 입력한 뒤 저장하세요.")

        mc1, mc2, mc3 = st.columns(3)
        with mc1:
            m_date    = st.text_input("날짜 (YYYY-MM-DD)", placeholder="2026-03-01", key="manual_date")
        with mc2:
            m_media   = st.selectbox("매체", [""] + list(COMMENT_MEDIA_MAP.keys()), key="manual_media_sel")
        with mc3:
            m_landing = st.text_input("랜딩 (선택)", placeholder="올리브영", key="manual_landing")

        m_comment = st.text_area(
            "코멘트 내용 붙여넣기",
            height=200,
            placeholder="* Meta — 올리브영\n- 3/1(일) 광고비 약 237만 원 소진...\nㄴ 전주 대비 CVR 급증...",
            key="manual_comment",
        )

        if st.button("💾 저장", type="primary", use_container_width=True, key="manual_save"):
            if not m_date or not m_media or not m_comment.strip():
                st.warning("날짜, 매체, 코멘트 내용을 모두 입력하세요.")
            else:
                entry = {
                    "date":    m_date.strip(),
                    "media":   m_media.strip(),
                    "landing": m_landing.strip(),
                    "comment": m_comment.strip(),
                }
                st.session_state.comment_history.append(entry)
                st.session_state.comment_history.sort(key=lambda x: x.get("date",""))
                st.success(f"✅ 저장 완료 — {m_date} / {m_media}")
                st.rerun()

    # ════════════════════════════════════════════════════════
    # TAB 3 — 저장된 히스토리 보기 & 관리
    # ════════════════════════════════════════════════════════
    with tab_view:
        st.subheader("저장된 코멘트 히스토리")
        history = st.session_state.comment_history

        if not history:
            st.info("저장된 히스토리가 없습니다. '파일 업로드' 또는 '직접 입력' 탭에서 저장하세요.")
        else:
            # 필터
            fc1, fc2 = st.columns(2)
            with fc1:
                all_media_opts = ["전체"] + sorted(set(h.get("media","") for h in history if h.get("media")))
                filter_media   = st.selectbox("매체 필터", all_media_opts, key="hist_filter_media")
            with fc2:
                filter_kw = st.text_input("코멘트 키워드 검색", placeholder="CVR, ROAS, 소재명 등", key="hist_kw")

            filtered = [
                h for h in history
                if (filter_media == "전체" or h.get("media","") == filter_media)
                and (not filter_kw or filter_kw.lower() in h.get("comment","").lower())
            ]

            st.caption(f"전체 {len(history)}건 중 {len(filtered)}건 표시")

            # 목록 표시
            for i, h in enumerate(reversed(filtered)):  # 최신순
                with st.expander(
                    f"📅 {h.get('date','-')} | {h.get('media','-')}"
                    + (f" / {h['landing']}" if h.get('landing') else ""),
                    expanded=False,
                ):
                    st.text(h.get("comment",""))
                    if st.button("🗑️ 이 항목 삭제", key=f"del_{i}_{h.get('date')}"):
                        orig_idx = len(history) - 1 - filtered.index(h) if h in filtered else -1
                        try:
                            st.session_state.comment_history.remove(h)
                            st.rerun()
                        except ValueError:
                            pass

            st.divider()

            # 전체 내보내기
            st.caption("**히스토리 내보내기**")
            df_export = pd.DataFrame(history)[["date","media","landing","comment"]]
            dcol1, dcol2 = st.columns(2)
            with dcol1:
                # txt 형식으로 내보내기
                txt_out = ""
                for h in history:
                    landing_str = f" / {h['landing']}" if h.get("landing") else ""
                    txt_out += f"[{h.get('date','')} / {h.get('media','')}{landing_str}]\n"
                    txt_out += h.get("comment","").strip() + "\n\n"
                st.download_button(
                    "⬇️ txt로 내보내기",
                    data=txt_out.encode("utf-8-sig"),
                    file_name=f"comment_history_{datetime.today().strftime('%Y%m%d')}.txt",
                    mime="text/plain", use_container_width=True,
                )
            with dcol2:
                st.download_button(
                    "⬇️ csv로 내보내기",
                    data=df_export.to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"comment_history_{datetime.today().strftime('%Y%m%d')}.csv",
                    mime="text/csv", use_container_width=True,
                )
